"""上传经验 Common_Chart 的校验、求值与匹配测试。"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from stds.domain.models import MostChart, ValueOption
from stds.experience import (
    CommonChartEntry,
    CommonChartKind,
    load_experience_workbook,
    match_common_chart,
    normalize_common_keyword,
)


def _option(
    variable: int,
    value_number: int,
    abbreviation: str,
    formula_value: float,
    next_variable: int = 0,
) -> ValueOption:
    return ValueOption(
        variable_number=variable,
        range_number=1,
        value_number=value_number,
        description=abbreviation,
        metric_abbrev=abbreviation,
        formula_value=formula_value,
        next_variable=next_variable,
        next_range=1 if next_variable else 0,
    )


def _charts() -> dict[str, MostChart]:
    return {
        "050 222": MostChart(
            chartcode="050 222",
            title="Move",
            formula="V2",
            value_added=False,
            developed_in_seconds=True,
            options={
                (1, 1): [_option(1, 1, "UOBS", 0.48, 2)],
                (2, 1): [
                    # 数值退化会错误命中这一项；全候选精确匹配应选下一项。
                    _option(2, 1, "2.0MX", 3.0),
                    _option(2, 2, "3.0MX", 4.0),
                ],
            },
        ),
        "050 221": MostChart(
            chartcode="050 221",
            title="Get",
            formula="V2",
            value_added=False,
            developed_in_seconds=True,
            options={
                (1, 1): [_option(1, 1, "SIM", 1.0, 2)],
                (2, 1): [_option(2, 1, "NARX", 1.0)],
            },
        ),
        "EST C00": MostChart(
            chartcode="EST C00",
            title="Fixed C",
            formula="V1",
            value_added=True,
            developed_in_seconds=True,
            options={},
        ),
        "EST V00": MostChart(
            chartcode="EST V00",
            title="Fixed V",
            formula="V1",
            value_added=False,
            developed_in_seconds=True,
            options={},
        ),
    }


def _workbook_bytes(common_rows: list[tuple]) -> bytes:
    workbook = Workbook()
    chart_sheet = workbook.active
    chart_sheet.title = "chartcode选择经验"
    chart_sheet.append(["操作内容", "参数选择"])
    parameter_sheet = workbook.create_sheet("参数选择经验")
    parameter_sheet.append(["操作内容", "动作代码", "参数选择经验"])
    common_sheet = workbook.create_sheet("Common_Chart")
    common_sheet.append([
        "序号",
        "操作内容",
        "决策描述",
        "动作代码",
        "增值/非增值(C/V)",
        "频率",
        "时间",
        "关键词描述1",
        "关键词描述2",
        "关键词描述3",
        "关键词描述4",
        "关键词描述5",
        "关键词描述6",
        "关键词描述7",
        "关键词描述8",
    ])
    for row in common_rows:
        common_sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _row(
    number: int,
    operation: str,
    decision: str | None,
    chartcode: str,
    cv: str,
    frequency: object,
    time_s: object,
    *keywords: str,
) -> tuple:
    return (
        number,
        operation,
        decision,
        chartcode,
        cv,
        frequency,
        time_s,
        *keywords,
    )


def test_loader_recalculates_formula_canonicalizes_alias_and_loads_fixed_est():
    source = _workbook_bytes([
        _row(1, "行走3米", "UOBS,3.0MX", "050 222", "V", 1, 3, "行走3米"),
        _row(2, "拿取", "SIM,NAR", "050 221", "V", 1, 1, "拿取零件"),
        _row(3, "拧紧", None, "EST C00", "C", 1, 5, "拧紧", "预紧"),
    ])

    result = load_experience_workbook(source, _charts())

    assert not result.has_errors
    assert len(result.common_entries) == 3
    move, get, fixed = result.common_entries
    assert move.kind is CommonChartKind.FORMULA
    assert move.values == {1: 0.48, 2: 4.0}
    assert move.source_time_s == 3.0
    assert move.time_s == 4.0
    assert get.decision == "SIM,NARX"
    assert get.values[2] == 1.0
    assert fixed.kind is CommonChartKind.FIXED_TIME
    assert fixed.chartcode == "EST C00"
    assert fixed.decision == "5S"
    assert fixed.values == {}
    assert fixed.time_s == 5.0
    assert any(
        issue.code == "common_time_mismatch" and issue.row == 2
        for issue in result.issues
    )


def test_loader_disables_only_invalid_common_rows():
    source = _workbook_bytes([
        _row(1, "有效", None, "EST V00", "V", 1, 2, "有效动作"),
        _row(2, "频率错误", None, "EST V00", "V", 4, 8, "频率错误"),
        _row(3, "CV错误", None, "EST C00", "V", 1, 5, "属性错误"),
        _row(4, "代码错误", None, "EST X00", "V", 1, 5, "代码错误"),
        _row(5, "决策错误", "UNKNOWN", "050 222", "V", 1, 5, "决策错误"),
        _row(6, "时间错误", None, "EST V00", "V", 1, "bad", "时间错误"),
        _row(7, "关键词错误", None, "EST V00", "V", 1, 2, "拿"),
    ])

    result = load_experience_workbook(source, _charts())

    assert [entry.operation_label for entry in result.common_entries] == [
        "有效",
        "关键词错误",
    ]
    semantic_only = result.common_entries[1]
    assert semantic_only.keywords == ()
    assert semantic_only.normalized_keywords == ()
    assert {
        issue.code
        for issue in result.issues
    }.issuperset({
        "invalid_common_frequency",
        "common_cv_mismatch",
        "invalid_common_chartcode",
        "invalid_common_decision",
        "invalid_common_time",
        "invalid_common_keyword",
        "no_valid_common_keyword",
    })


def _entry(
    row: int,
    keyword: str,
    *,
    chartcode: str = "EST V00",
    time_s: float = 1.0,
) -> CommonChartEntry:
    return CommonChartEntry(
        operation_label=keyword,
        normalized_operation=normalize_common_keyword(keyword),
        chartcode=chartcode,
        decision=f"{time_s:g}S",
        cv="V",
        frequency=1.0,
        source_time_s=time_s,
        time_s=time_s,
        keywords=(keyword,),
        normalized_keywords=(normalize_common_keyword(keyword),),
        row=row,
        kind=CommonChartKind.FIXED_TIME,
        values={},
    )


def test_matcher_uses_exact_then_longest_and_never_reverse_contains():
    entries = (
        _entry(2, "转身", time_s=1),
        _entry(3, "转身90", time_s=2),
        _entry(4, "拿取", time_s=3),
    )

    exact = match_common_chart("转身", entries)
    longest = match_common_chart("人工A转身90度", entries)

    assert exact is not None
    assert exact.entry.row == 2
    assert exact.match_type == "exact"
    assert longest is not None
    assert longest.entry.row == 3
    assert longest.keyword == "转身90"
    assert longest.match_type == "contains"
    assert match_common_chart("拿", entries) is None


def test_matcher_falls_back_on_equal_length_different_outputs():
    entries = (
        _entry(2, "扫描", time_s=1),
        _entry(3, "扫描", time_s=2),
    )

    assert match_common_chart("人工A扫描", entries) is None


def test_matcher_deduplicates_equivalent_outputs_deterministically():
    entries = (
        _entry(8, "扫描", time_s=1.2),
        _entry(3, "扫描", time_s=1.2),
    )

    match = match_common_chart("扫描", entries)

    assert match is not None
    assert match.entry.row == 3


def test_keyword_minimum_length_supports_cjk_ascii_and_mixed_units():
    assert normalize_common_keyword("拿") == ""
    assert normalize_common_keyword("拿取") == "拿取"
    assert normalize_common_keyword("ab") == ""
    assert normalize_common_keyword("abc") == "abc"
    assert normalize_common_keyword("走1m") == "走1m"
