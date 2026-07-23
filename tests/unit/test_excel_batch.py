"""STDS PF 七列输入、八列拆解和十四列工时结果契约。"""
from __future__ import annotations

import asyncio
import csv
from io import BytesIO, StringIO

import pytest
from openpyxl import Workbook, load_workbook

from stds.domain.models import Source, StdsResult
from stds.pipeline.excel_batch import (
    DECOMPOSITION_HEADERS,
    INPUT_HEADERS,
    INPUT_SHEET_NAME,
    OUTPUT_HEADERS,
    ExcelInputError,
    analyze_decomposition_output,
    analyze_excel_bytes,
    decompose_excel_bytes,
    review_decomposition_rows,
)


EXPECTED_INPUT_HEADERS = (
    "序号",
    "项目名称",
    "产品型号",
    "产线",
    "工位号",
    "工位描述",
    "作业描述",
)
EXPECTED_DECOMPOSITION_HEADERS = (
    *EXPECTED_INPUT_HEADERS,
    "翻译后作业描述",
)
EXPECTED_OUTPUT_HEADERS = (
    "序号",
    "项目名称",
    "产品型号",
    "产线",
    "工位号",
    "工位描述",
    "SOS描述",
    "JES描述",
    "STDS描述",
    "Decisions",
    "Chart",
    "增值|非增值",
    "Freq",
    "Time(s)",
)
DEFAULT_ROW = [
    88,
    "项目A",
    "Model-X",
    "Line-1",
    "OP010",
    "装配工位",
    "人工拿取零件",
]


async def _passthrough_decomposer(operation: str) -> list[str]:
    return [operation]


async def _analyze_excel_bytes(*args, **kwargs):
    kwargs.setdefault("decomposer", _passthrough_decomposer)
    return await analyze_excel_bytes(*args, **kwargs)


def _workbook_bytes(
    rows: list[list] | None = None,
    *,
    headers: tuple[str, ...] = EXPECTED_INPUT_HEADERS,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = INPUT_SHEET_NAME
    ws.append(list(headers))
    for row in rows or [DEFAULT_ROW]:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _result(element, *, decision: str = "T,90,NB") -> StdsResult:
    return StdsResult(
        element=element,
        chartcode="202 010",
        decision=decision,
        time_s=1.2,
        cv="V",
        freq=1.0,
        source=Source.FORMULA,
        confidence=0.9,
        needs_review=False,
        trace=[("V1", "Turn", "operation-match")],
    )


def _sheet(payload: bytes):
    return load_workbook(BytesIO(payload), data_only=False)[INPUT_SHEET_NAME]


def _headers(ws) -> tuple:
    return tuple(ws.cell(1, col).value for col in range(1, ws.max_column + 1))


def _csv_rows(payload: bytes) -> list[list[str]]:
    assert payload.startswith(b"\xef\xbb\xbf")
    return list(csv.reader(StringIO(payload.decode("utf-8-sig"))))


def _sheet_display_rows(ws) -> list[list[str]]:
    rows = []
    for row_index in range(1, ws.max_row + 1):
        values = []
        for col_index in range(1, ws.max_column + 1):
            cell = ws.cell(row_index, col_index)
            value = cell.value
            if value is None:
                values.append("")
            elif cell.number_format == "0.00" and isinstance(value, (int, float)):
                values.append(f"{value:.2f}")
            elif cell.number_format == "0.##" and isinstance(value, (int, float)):
                values.append(f"{value:.2f}".rstrip("0").rstrip("."))
            elif isinstance(value, float) and value.is_integer():
                values.append(str(int(value)))
            else:
                values.append(str(value))
        rows.append(values)
    return rows


def test_public_headers_match_the_three_pf_templates():
    assert INPUT_HEADERS == EXPECTED_INPUT_HEADERS
    assert DECOMPOSITION_HEADERS == EXPECTED_DECOMPOSITION_HEADERS
    assert OUTPUT_HEADERS == EXPECTED_OUTPUT_HEADERS


def test_manual_review_first_stage_only_decomposes_and_translates(monkeypatch):
    resolver_calls = 0
    original_operations = ["Manual pick part", "操作人员检查零件"]

    async def forbidden_resolve_with_actor(*args, **kwargs):
        nonlocal resolver_calls
        resolver_calls += 1
        raise AssertionError("阶段一不应调用工时 resolver")

    async def fake_decomposer(operation):
        assert operation == "Manual assemble module"
        return original_operations

    async def fake_translator(operation):
        assert operation == "Manual pick part"
        return "操作人员拿取零件"

    monkeypatch.setattr(
        "stds.pipeline.excel_batch.resolve_with_actor",
        forbidden_resolve_with_actor,
    )
    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(
                [[1, "项目A", "M1", "L1", "OP010", "装配", "Manual assemble module"]]
            ),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=fake_decomposer,
            translator=fake_translator,
        )
    )

    assert resolver_calls == 0
    assert [timing.phase for timing in stage.timings] == ["拆解"]
    assert [row["作业描述"] for row in stage.decomposition_rows()] == (
        original_operations
    )
    assert [row["翻译后作业描述"] for row in stage.decomposition_rows()] == [
        "操作人员拿取零件",
        "操作人员检查零件",
    ]
    decomposition_ws = _sheet(stage.decomposition_bytes)
    assert _headers(decomposition_ws) == EXPECTED_DECOMPOSITION_HEADERS
    assert decomposition_ws.max_column == 8
    assert decomposition_ws.max_row == 3


def test_reviewed_rows_support_edit_add_delete_reorder_and_split_g_h_usage():
    async def fake_decomposer(operation):
        assert operation == "操作人员装配模块"
        return ["原动作甲", "原动作乙", "原动作丙"]

    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(
                [[1, "项目A", "M1", "L1", "OP010", "装配", "操作人员装配模块"]]
            ),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=fake_decomposer,
        )
    )
    original_rows = stage.decomposition_rows()

    edited_third = dict(original_rows[2])
    edited_third.update(
        {
            "序号": 300,
            "作业描述": "Manual reviewed operation",
            "翻译后作业描述": "人工审核后的翻译",
        }
    )
    added_row = {
        "序号": 999,
        "项目名称": "新增项目",
        "产品型号": "M2",
        "产线": "L2",
        "工位号": "OP020",
        "工位描述": "新增工位",
        "作业描述": "新增审核动作",
        "翻译后作业描述": "新增审核翻译",
    }
    retained_first = dict(original_rows[0])
    retained_first["序号"] = -1

    reviewed = review_decomposition_rows(
        stage,
        [edited_third, added_row, retained_first],
    )
    expected_originals = [
        "Manual reviewed operation",
        "新增审核动作",
        "原动作甲",
    ]
    expected_translations = [
        "人工审核后的翻译",
        "新增审核翻译",
        "原动作甲",
    ]
    reviewed_rows = reviewed.decomposition_rows()
    assert [row["序号"] for row in reviewed_rows] == [1, 2, 3]
    assert [row["作业描述"] for row in reviewed_rows] == expected_originals
    assert [row["翻译后作业描述"] for row in reviewed_rows] == (
        expected_translations
    )
    assert "原动作乙" not in [row["作业描述"] for row in reviewed_rows]
    reviewed_ws = _sheet(reviewed.decomposition_bytes)
    assert _headers(reviewed_ws) == EXPECTED_DECOMPOSITION_HEADERS
    assert reviewed_ws.max_column == 8
    assert [reviewed_ws.cell(row, 1).value for row in range(2, 5)] == [1, 2, 3]
    assert reviewed.decomposition_csv_filename == "2.STDS-PF拆解.csv"
    assert _csv_rows(reviewed.decomposition_csv_bytes) == _sheet_display_rows(
        reviewed_ws
    )
    assert reviewed.decomposition_display_rows() == [
        dict(zip(EXPECTED_DECOMPOSITION_HEADERS, values))
        for values in _csv_rows(reviewed.decomposition_csv_bytes)[1:]
    ]

    analyzed_inputs = []

    async def fake_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is False
        analyzed_inputs.append(element.operation_des)
        return _result(element)

    batch = asyncio.run(
        analyze_decomposition_output(
            reviewed,
            object(),
            resolver=fake_resolver,
        )
    )

    assert analyzed_inputs == expected_originals
    assert [row["STDS描述"] for row in batch.detail_preview_rows()] == (
        expected_translations
    )
    final_ws = _sheet(batch.output_bytes)
    assert _headers(final_ws) == EXPECTED_OUTPUT_HEADERS
    assert final_ws.max_column == 14
    assert [final_ws.cell(row, 1).value for row in range(2, 5)] == [1, 2, 3]
    assert [final_ws.cell(row, 9).value for row in range(2, 5)] == (
        expected_translations
    )
    assert all(
        original not in [final_ws.cell(row, 9).value for row in range(2, 5)]
        for original in expected_originals[:2]
    )
    assert batch.output_csv_filename == "3.STDS-工时生成.csv"
    assert _csv_rows(batch.output_csv_bytes) == _sheet_display_rows(final_ws)
    assert batch.detail_display_rows() == [
        dict(zip(EXPECTED_OUTPUT_HEADERS, values))
        for values in _csv_rows(batch.output_csv_bytes)[1:]
    ]


def test_reviewed_rows_reject_empty_operation_or_empty_table():
    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=_passthrough_decomposer,
        )
    )

    empty_operation = dict(stage.decomposition_rows()[0])
    empty_operation["作业描述"] = "  "
    with pytest.raises(ExcelInputError, match="作业描述"):
        review_decomposition_rows(stage, [empty_operation])

    empty_translation = dict(stage.decomposition_rows()[0])
    empty_translation["翻译后作业描述"] = "  "
    with pytest.raises(ExcelInputError, match="翻译后作业描述"):
        review_decomposition_rows(stage, [empty_translation])

    with pytest.raises(ExcelInputError):
        review_decomposition_rows(stage, [])

    blank_editor_row = {header: None for header in EXPECTED_DECOMPOSITION_HEADERS}
    reviewed = review_decomposition_rows(
        stage,
        [blank_editor_row, stage.decomposition_rows()[0]],
    )
    assert reviewed.detail_count == 1


def test_reviewed_formula_like_text_stays_literal_in_download():
    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=_passthrough_decomposer,
        )
    )
    reviewed_row = dict(stage.decomposition_rows()[0])
    reviewed_row["作业描述"] = "=1+1"
    reviewed_row["翻译后作业描述"] = "=人工确认文本"
    reviewed = review_decomposition_rows(stage, [reviewed_row])
    ws = _sheet(reviewed.decomposition_bytes)

    assert ws.cell(2, 7).value == "=1+1"
    assert ws.cell(2, 7).data_type == "s"
    assert ws.cell(2, 8).value == "=人工确认文本"
    assert ws.cell(2, 8).data_type == "s"
    csv_rows = _csv_rows(reviewed.decomposition_csv_bytes)
    assert csv_rows[1][6] == "'=1+1"
    assert csv_rows[1][7] == "'=人工确认文本"


def test_reviewed_auto_operation_is_reclassified_as_machine_without_redecomposition():
    decomposer_calls = 0

    async def original_decomposer(operation):
        nonlocal decomposer_calls
        decomposer_calls += 1
        return [operation]

    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=original_decomposer,
        )
    )
    reviewed_row = dict(stage.decomposition_rows()[0])
    reviewed_row["作业描述"] = "Auto Robot Load CTR to pallet"
    reviewed_row["翻译后作业描述"] = "自动 Robot 装载 CTR 到 pallet"
    reviewed = review_decomposition_rows(stage, [reviewed_row])

    async def machine_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is True
        return StdsResult.machine_placeholder(element)

    batch = asyncio.run(
        analyze_decomposition_output(
            reviewed,
            object(),
            resolver=machine_resolver,
        )
    )

    assert decomposer_calls == 1
    assert reviewed.rows[0].split.actor == "设备"
    assert batch.detail_preview_rows()[0]["STDS描述"] == (
        "自动 Robot 装载 CTR 到 pallet"
    )
    assert list(batch.detail_preview_rows()[0].values())[9:14] == ["NA"] * 5


def test_reviewed_ambiguous_operation_uses_llm_actor_without_redecomposition():
    decomposer_calls = 0
    classifier_calls = 0

    async def original_decomposer(operation):
        nonlocal decomposer_calls
        decomposer_calls += 1
        return [operation]

    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=original_decomposer,
        )
    )
    reviewed_row = dict(stage.decomposition_rows()[0])
    reviewed_row["作业描述"] = "机械手完成上件"
    reviewed_row["翻译后作业描述"] = "机械手完成上件"
    reviewed = review_decomposition_rows(stage, [reviewed_row])

    class ClassifierDeps:
        async def llm_classify(self, operation):
            nonlocal classifier_calls
            classifier_calls += 1
            assert operation == "机械手完成上件"
            return True

    async def machine_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is True
        return StdsResult.machine_placeholder(element)

    batch = asyncio.run(
        analyze_decomposition_output(
            reviewed,
            ClassifierDeps(),
            resolver=machine_resolver,
        )
    )

    assert decomposer_calls == 1
    assert classifier_calls == 1
    assert batch.rows[0].split.actor == "设备"
    assert list(batch.detail_preview_rows()[0].values())[9:14] == ["NA"] * 5


def test_reviewed_unchanged_operation_reuses_llm_actor_even_if_number_changes():
    classifier_calls = 0

    class ClassifierDeps:
        async def llm_classify(self, operation):
            nonlocal classifier_calls
            classifier_calls += 1
            assert operation == "机械手完成上件"
            return True

    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(
                [[1, "项目A", "M1", "L1", "OP010", "自动线", "机械手完成上件"]]
            ),
            "1.STDS-PF清单.xlsx",
            ClassifierDeps(),
            decomposer=_passthrough_decomposer,
        )
    )
    reviewed_row = dict(stage.decomposition_rows()[0])
    reviewed_row["序号"] = 999
    reviewed = review_decomposition_rows(stage, [reviewed_row])

    async def machine_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is True
        return StdsResult.machine_placeholder(element)

    batch = asyncio.run(
        analyze_decomposition_output(
            reviewed,
            object(),
            resolver=machine_resolver,
        )
    )

    assert classifier_calls == 1
    assert reviewed.rows[0].split.actor == "设备"
    assert batch.rows[0].split.actor == "设备"


def test_reviewed_actor_classification_failure_is_marked_for_review():
    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=_passthrough_decomposer,
        )
    )
    reviewed_row = dict(stage.decomposition_rows()[0])
    reviewed_row["作业描述"] = "机械手完成上件"
    reviewed_row["翻译后作业描述"] = "机械手完成上件"
    reviewed = review_decomposition_rows(stage, [reviewed_row])

    class BrokenClassifierDeps:
        async def llm_classify(self, operation):
            raise RuntimeError("classifier unavailable")

    async def fake_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is False
        return _result(element)

    batch = asyncio.run(
        analyze_decomposition_output(
            reviewed,
            BrokenClassifierDeps(),
            resolver=fake_resolver,
        )
    )

    assert batch.review_count == 1
    assert batch.rows[0].split.needs_review is True
    assert batch.detail_preview_rows()[0]["Time(s)"] == "NA"


def test_review_reclassifies_an_untrusted_initial_actor_after_recovery():
    class BrokenClassifierDeps:
        async def llm_classify(self, operation):
            raise RuntimeError("temporary classifier failure")

    stage = asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(
                [[1, "项目A", "M1", "L1", "OP010", "自动线", "机械手完成上件"]]
            ),
            "1.STDS-PF清单.xlsx",
            BrokenClassifierDeps(),
            decomposer=_passthrough_decomposer,
        )
    )
    assert stage.rows[0].split.needs_review is True
    reviewed = review_decomposition_rows(stage, stage.decomposition_rows())

    class RecoveredClassifierDeps:
        async def llm_classify(self, operation):
            return True

    async def machine_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is True
        return StdsResult.machine_placeholder(element)

    batch = asyncio.run(
        analyze_decomposition_output(
            reviewed,
            RecoveredClassifierDeps(),
            resolver=machine_resolver,
        )
    )

    assert batch.rows[0].split.actor == "设备"
    assert batch.rows[0].split.needs_review is False
    assert batch.review_count == 0


def test_pf_input_is_flattened_to_independent_decomposition_and_final_files():
    source_metadata = ["项目A", "Model-X", "Line-1", "OP010", "装配工位"]
    child_operations = [
        "操作人员拿取吊具",
        "操作人员移动吊具",
        "操作人员使用吊具抓取中心支撑壳体",
        "操作人员移动吊具",
        "操作人员操作吊具释放中心支撑壳体",
    ]

    async def fake_decomposer(operation):
        assert operation == "操作人员用吊具转运中心支撑壳体"
        return child_operations

    async def fake_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is False
        assert element.operation_des in child_operations
        return _result(element)

    batch = asyncio.run(
        analyze_excel_bytes(
            _workbook_bytes(
                [
                    [
                        20,
                        *source_metadata,
                        "操作人员用吊具转运中心支撑壳体",
                    ]
                ]
            ),
            "1.STDS-PF清单 副本.xlsx",
            object(),
            resolver=fake_resolver,
            decomposer=fake_decomposer,
        )
    )

    assert batch.decomposition_filename == "2.STDS-PF拆解 副本.xlsx"
    assert batch.output_filename == "3.STDS-工时生成 副本.xlsx"
    assert batch.decomposition_bytes != batch.output_bytes

    decomposition_ws = _sheet(batch.decomposition_bytes)
    assert _headers(decomposition_ws) == EXPECTED_DECOMPOSITION_HEADERS
    assert decomposition_ws.max_column == 8
    assert decomposition_ws.max_row == 6
    assert [decomposition_ws.cell(row, 1).value for row in range(2, 7)] == list(
        range(1, 6)
    )
    for column, value in enumerate(source_metadata, start=2):
        assert [decomposition_ws.cell(row, column).value for row in range(2, 7)] == [
            value
        ] * 5
    assert [decomposition_ws.cell(row, 7).value for row in range(2, 7)] == (
        child_operations
    )
    assert [decomposition_ws.cell(row, 8).value for row in range(2, 7)] == (
        child_operations
    )
    assert decomposition_ws.freeze_panes is None
    assert decomposition_ws.auto_filter.ref is None

    final_ws = _sheet(batch.output_bytes)
    assert _headers(final_ws) == EXPECTED_OUTPUT_HEADERS
    assert final_ws.max_column == 14
    assert final_ws.max_row == 6
    assert [final_ws.cell(row, 1).value for row in range(2, 7)] == list(range(1, 6))
    for column, value in enumerate(source_metadata, start=2):
        assert [final_ws.cell(row, column).value for row in range(2, 7)] == [value] * 5
    assert [final_ws.cell(row, 7).value for row in range(2, 7)] == ["NA"] * 5
    assert [final_ws.cell(row, 8).value for row in range(2, 7)] == ["NA"] * 5
    assert [final_ws.cell(row, 9).value for row in range(2, 7)] == child_operations
    assert [final_ws.cell(row, 10).value for row in range(2, 7)] == [
        "T,90,NB"
    ] * 5
    assert [final_ws.cell(row, 11).value for row in range(2, 7)] == [
        "202 010"
    ] * 5
    assert [final_ws.cell(row, 12).value for row in range(2, 7)] == ["V"] * 5
    assert [final_ws.cell(row, 13).value for row in range(2, 7)] == [1.0] * 5
    assert [final_ws.cell(row, 14).value for row in range(2, 7)] == [1.2] * 5
    assert all(final_ws.cell(row, 13).number_format == "0.##" for row in range(2, 7))
    assert all(final_ws.cell(row, 14).number_format == "0.00" for row in range(2, 7))
    assert final_ws.freeze_panes is None
    assert final_ws.auto_filter.ref is None
    assert all(
        final_ws.cell(row, col).value is None
        for row in range(1, 7)
        for col in range(15, 18)
    )

    decomposition_rows = batch.decomposition_rows()
    assert tuple(decomposition_rows[0]) == EXPECTED_DECOMPOSITION_HEADERS
    assert [row["序号"] for row in decomposition_rows] == list(range(1, 6))
    assert [row["作业描述"] for row in decomposition_rows] == child_operations
    assert [row["翻译后作业描述"] for row in decomposition_rows] == child_operations

    detail_rows = batch.detail_preview_rows()
    assert tuple(detail_rows[0]) == EXPECTED_OUTPUT_HEADERS
    assert [row["序号"] for row in detail_rows] == list(range(1, 6))
    assert [row["STDS描述"] for row in detail_rows] == child_operations


def test_global_sequence_is_rebuilt_and_duplicate_operations_are_analyzed_once():
    resolver_calls = 0
    progress = []

    async def fake_resolver(element, deps, *, machine_hint=None):
        nonlocal resolver_calls
        resolver_calls += 1
        return _result(element)

    batch = asyncio.run(
        _analyze_excel_bytes(
            _workbook_bytes(
                [
                    [10, "项目甲", "M1", "L1", "OP010", "前工位", "重复操作"],
                    [20, "项目乙", "M2", "L2", "OP020", "后工位", "重复操作"],
                ]
            ),
            "重复.xlsx",
            object(),
            resolver=fake_resolver,
            on_progress=progress.append,
        )
    )

    assert resolver_calls == 1
    assert batch.total_count == 2
    assert batch.detail_count == 2
    assert len(batch.timing_rows()) == 4
    assert all(item.item_elapsed_s >= 0 for item in progress)
    assert progress[-1].overall_ratio == 1.0
    assert batch.decomposition_filename == "重复_PF拆解.xlsx"
    assert batch.output_filename == "重复_工时生成.xlsx"

    decomposition_ws = _sheet(batch.decomposition_bytes)
    final_ws = _sheet(batch.output_bytes)
    assert [decomposition_ws.cell(row, 1).value for row in (2, 3)] == [1, 2]
    assert [final_ws.cell(row, 1).value for row in (2, 3)] == [1, 2]
    assert [decomposition_ws.cell(row, 2).value for row in (2, 3)] == [
        "项目甲",
        "项目乙",
    ]
    assert [final_ws.cell(row, 5).value for row in (2, 3)] == ["OP010", "OP020"]


@pytest.mark.parametrize(
    ("source_name", "decomposition_name", "output_name"),
    [
        (
            "1.STDS-PF清单.xlsx",
            "2.STDS-PF拆解.xlsx",
            "3.STDS-工时生成.xlsx",
        ),
        (
            "1.STDS-PF清单-v2.xlsx",
            "2.STDS-PF拆解-v2.xlsx",
            "3.STDS-工时生成-v2.xlsx",
        ),
        (
            "/tmp/普通文件.xlsx",
            "普通文件_PF拆解.xlsx",
            "普通文件_工时生成.xlsx",
        ),
    ],
)
def test_output_filenames_follow_pf_naming_contract(
    source_name,
    decomposition_name,
    output_name,
):
    async def fake_resolver(element, deps, *, machine_hint=None):
        return _result(element)

    batch = asyncio.run(
        _analyze_excel_bytes(
            _workbook_bytes(),
            source_name,
            object(),
            resolver=fake_resolver,
        )
    )

    assert batch.decomposition_filename == decomposition_name
    assert batch.output_filename == output_name


def test_only_final_stds_description_is_translated_analysis_uses_original_text():
    translated_inputs = []
    analyzed_inputs = []
    original_operations = [
        "Manual pick up Front End Module",
        "操作人员 install ECU bracket",
        "操作人员拿取中文零件",
    ]

    async def fake_decomposer(operation):
        assert operation == "Manual assemble module"
        return original_operations

    async def fake_translator(operation):
        translated_inputs.append(operation)
        return {
            "Manual pick up Front End Module": "操作人员拿取 Front End Module",
            "操作人员 install ECU bracket": "操作人员安装 ECU bracket",
        }[operation]

    async def fake_resolver(element, deps, *, machine_hint=None):
        analyzed_inputs.append(element.operation_des)
        return _result(element)

    batch = asyncio.run(
        analyze_excel_bytes(
            _workbook_bytes(
                [[1, "项目A", "M1", "L1", "OP010", "装配", "Manual assemble module"]]
            ),
            "中英混合.xlsx",
            object(),
            resolver=fake_resolver,
            decomposer=fake_decomposer,
            translator=fake_translator,
        )
    )

    expected_output = [
        "操作人员拿取 Front End Module",
        "操作人员安装 ECU bracket",
        "操作人员拿取中文零件",
    ]
    assert analyzed_inputs == original_operations
    assert sorted(translated_inputs) == sorted(original_operations[:2])
    assert [detail.operation for detail in batch.rows[0].details] == original_operations
    decomposition_rows = batch.decomposition_rows()
    assert [row["作业描述"] for row in decomposition_rows] == original_operations
    assert [row["翻译后作业描述"] for row in decomposition_rows] == expected_output
    assert [row["STDS描述"] for row in batch.detail_preview_rows()] == expected_output

    decomposition_ws = _sheet(batch.decomposition_bytes)
    assert _headers(decomposition_ws) == EXPECTED_DECOMPOSITION_HEADERS
    assert decomposition_ws.max_column == 8
    assert [decomposition_ws.cell(row, 7).value for row in range(2, 5)] == (
        original_operations
    )
    assert [decomposition_ws.cell(row, 8).value for row in range(2, 5)] == (
        expected_output
    )

    final_ws = _sheet(batch.output_bytes)
    assert _headers(final_ws) == EXPECTED_OUTPUT_HEADERS
    assert final_ws.max_column == 14
    assert [final_ws.cell(row, 9).value for row in range(2, 5)] == expected_output
    assert all(final_ws.cell(row, 7).value == "NA" for row in range(2, 5))
    assert all(final_ws.cell(row, 8).value == "NA" for row in range(2, 5))
    assert "作业描述" not in _headers(final_ws)
    assert "翻译后作业描述" not in _headers(final_ws)


def test_output_translation_is_deduplicated_and_failure_falls_back_to_original():
    translator_calls = 0

    async def fake_translator(operation):
        nonlocal translator_calls
        translator_calls += 1
        raise RuntimeError("translation service unavailable")

    async def fake_resolver(element, deps, *, machine_hint=None):
        return _result(element)

    batch = asyncio.run(
        _analyze_excel_bytes(
            _workbook_bytes(
                [
                    [1, "项目A", "M1", "L1", "OP010", "工位1", "Manual pick part"],
                    [2, "项目A", "M1", "L1", "OP020", "工位2", "Manual pick part"],
                ]
            ),
            "翻译回退.xlsx",
            object(),
            resolver=fake_resolver,
            translator=fake_translator,
        )
    )

    assert translator_calls == 1
    expected_fallback = ["Manual pick part", "Manual pick part"]
    decomposition_rows = batch.decomposition_rows()
    assert [row["作业描述"] for row in decomposition_rows] == expected_fallback
    assert [row["翻译后作业描述"] for row in decomposition_rows] == (
        expected_fallback
    )
    assert [row["STDS描述"] for row in batch.detail_preview_rows()] == [
        "Manual pick part",
        "Manual pick part",
    ]

    decomposition_ws = _sheet(batch.decomposition_bytes)
    assert _headers(decomposition_ws) == EXPECTED_DECOMPOSITION_HEADERS
    assert decomposition_ws.max_column == 8
    assert [decomposition_ws.cell(row, 7).value for row in (2, 3)] == (
        expected_fallback
    )
    assert [decomposition_ws.cell(row, 8).value for row in (2, 3)] == (
        expected_fallback
    )

    final_ws = _sheet(batch.output_bytes)
    assert _headers(final_ws) == EXPECTED_OUTPUT_HEADERS
    assert final_ws.max_column == 14
    assert [final_ws.cell(row, 9).value for row in (2, 3)] == expected_fallback


def test_auto_machine_translation_failure_still_uses_chinese_auto_prefix():
    async def fake_translator(operation):
        raise RuntimeError("translation service unavailable")

    async def machine_resolver(element, deps, *, machine_hint=None):
        assert machine_hint is True
        return StdsResult.machine_placeholder(element)

    batch = asyncio.run(
        _analyze_excel_bytes(
            _workbook_bytes(
                [[1, "项目A", "M1", "L1", "OP010", "自动线", "Auto Robot Load CTR to pallet"]]
            ),
            "设备英文.xlsx",
            object(),
            resolver=machine_resolver,
            translator=fake_translator,
        )
    )

    assert batch.decomposition_rows()[0]["作业描述"] == "Auto Robot Load CTR to pallet"
    assert batch.decomposition_rows()[0]["翻译后作业描述"] == (
        "自动 Robot Load CTR to pallet"
    )
    assert batch.detail_preview_rows()[0]["STDS描述"] == "自动 Robot Load CTR to pallet"
    final_ws = _sheet(batch.output_bytes)
    assert final_ws.cell(2, 9).value == "自动 Robot Load CTR to pallet"
    assert [final_ws.cell(2, col).value for col in range(10, 15)] == ["NA"] * 5


def test_unresolved_row_has_na_analysis_fields():
    async def unresolved_resolver(element, deps, *, machine_hint=None):
        return StdsResult.unresolved(element, None)

    batch = asyncio.run(
        _analyze_excel_bytes(
            _workbook_bytes(
                [[3, "项目A", "M1", "L1", "OP030", "未知工位", "无法识别的动作"]]
            ),
            "待复核.xlsx",
            object(),
            resolver=unresolved_resolver,
        )
    )

    assert batch.review_count == 1
    final_ws = _sheet(batch.output_bytes)
    assert final_ws.cell(2, 9).value == "无法识别的动作"
    assert [final_ws.cell(2, col).value for col in range(10, 15)] == ["NA"] * 5


def test_machine_operation_is_not_decomposed_and_outputs_na_analysis_fields():
    decomposer_calls = 0
    hints = []

    async def should_not_decompose(operation):
        nonlocal decomposer_calls
        decomposer_calls += 1
        return [operation]

    async def fake_resolver(element, deps, *, machine_hint=None):
        hints.append(machine_hint)
        return StdsResult.machine_placeholder(element)

    batch = asyncio.run(
        analyze_excel_bytes(
            _workbook_bytes(
                [[1, "项目A", "M1", "L1", "OP010", "自动线", "设备自动托盘进入"]]
            ),
            "设备.xlsx",
            object(),
            resolver=fake_resolver,
            decomposer=should_not_decompose,
        )
    )

    assert decomposer_calls == 0
    assert hints == [True]
    final_ws = _sheet(batch.output_bytes)
    assert final_ws.cell(2, 9).value == "设备自动托盘进入"
    assert [final_ws.cell(2, col).value for col in range(10, 15)] == ["NA"] * 5
    assert list(batch.detail_preview_rows()[0].values())[9:14] == ["NA"] * 5


def test_work_description_column_g_is_the_only_analysis_source():
    analyzed_inputs = []

    async def fake_resolver(element, deps, *, machine_hint=None):
        analyzed_inputs.append(element.operation_des)
        return _result(element)

    payload = _workbook_bytes(
        [[1, "项目A", "M1", "L1", "OP010", "装配工位", "读取G列", "不要读取H列"]],
        headers=(*EXPECTED_INPUT_HEADERS, "旧操作内容"),
    )
    batch = asyncio.run(
        _analyze_excel_bytes(payload, "读取G列.xlsx", object(), resolver=fake_resolver)
    )

    assert analyzed_inputs == ["读取G列"]
    assert batch.decomposition_rows()[0]["作业描述"] == "读取G列"
    assert batch.detail_preview_rows()[0]["STDS描述"] == "读取G列"


def test_old_three_column_template_is_rejected():
    payload = _workbook_bytes(
        [[1, "OP010", "人工拿取零件"]],
        headers=("序号", "工位号", "操作内容"),
    )

    with pytest.raises(ExcelInputError, match="项目名称"):
        asyncio.run(_analyze_excel_bytes(payload, "旧模板.xlsx", object()))


def test_header_hidden_characters_and_whitespace_are_cleaned():
    async def fake_resolver(element, deps, *, machine_hint=None):
        return _result(element)

    headers = (
        "\ufeff 序\n号 ",
        "\u3000项目 名称\t",
        "\u200b产品型号\u2060",
        "产 线",
        "工位\n号",
        "工位 描述",
        "作业\t描述",
    )
    batch = asyncio.run(
        _analyze_excel_bytes(
            _workbook_bytes(headers=headers),
            "隐藏字符模板.xlsx",
            object(),
            resolver=fake_resolver,
        )
    )

    decomposition_ws = _sheet(batch.decomposition_bytes)
    assert _headers(decomposition_ws) == EXPECTED_DECOMPOSITION_HEADERS
    assert decomposition_ws.max_column == 8
    assert decomposition_ws.cell(2, 2).value == "项目A"
    assert decomposition_ws.cell(2, 7).value == "人工拿取零件"
    assert decomposition_ws.cell(2, 8).value == "人工拿取零件"


def test_operation_formula_without_cached_value_is_rejected():
    wb = Workbook()
    ws = wb.active
    ws.title = INPUT_SHEET_NAME
    ws.append(list(EXPECTED_INPUT_HEADERS))
    ws.append([1, "项目A", "M1", "L1", "OP010", "装配工位", "=A2"])
    payload = BytesIO()
    wb.save(payload)

    with pytest.raises(ExcelInputError, match="没有已计算值"):
        asyncio.run(_analyze_excel_bytes(payload.getvalue(), "公式.xlsx", object()))


def test_empty_work_description_column_is_rejected():
    payload = _workbook_bytes(
        [[1, "项目A", "M1", "L1", "OP010", "装配工位", None]]
    )

    with pytest.raises(ExcelInputError, match="作业描述"):
        asyncio.run(_analyze_excel_bytes(payload, "空作业描述.xlsx", object()))


@pytest.mark.parametrize("payload", [b"", b"not-an-xlsx"])
def test_invalid_excel_has_user_friendly_error(payload):
    with pytest.raises(ExcelInputError):
        asyncio.run(_analyze_excel_bytes(payload, "bad.xlsx", object()))
