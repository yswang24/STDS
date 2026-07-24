"""线下审核版回传与审核表增删行的契约测试。"""
from __future__ import annotations

import asyncio
import csv
from io import BytesIO, StringIO

import pytest
from openpyxl import Workbook, load_workbook

from stds.pipeline.excel_batch import (
    DECOMPOSITION_HEADERS,
    INPUT_HEADERS,
    INPUT_SHEET_NAME,
    ExcelInputError,
    decompose_excel_bytes,
    parse_decomposition_review_upload,
    review_decomposition_rows,
)
from stds.ui.review_table import (
    delete_review_editor_rows,
    insert_review_editor_row,
    normalize_review_editor_rows,
)


def _workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = INPUT_SHEET_NAME
    worksheet.append(list(INPUT_HEADERS))
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def _passthrough_decomposer(operation: str) -> list[str]:
    return [operation]


async def _passthrough_translator(operation: str) -> str:
    return operation


def _decomposition_stage(rows: list[list[object]]):
    return asyncio.run(
        decompose_excel_bytes(
            _workbook_bytes(rows),
            "1.STDS-PF清单.xlsx",
            object(),
            decomposer=_passthrough_decomposer,
            translator=_passthrough_translator,
        )
    )


def _csv_rows(payload: bytes) -> list[list[str]]:
    assert payload.startswith(b"\xef\xbb\xbf")
    return list(csv.reader(StringIO(payload.decode("utf-8-sig"))))


def _metadata(row: dict[str, object]) -> tuple[object, ...]:
    return tuple(row[header] for header in DECOMPOSITION_HEADERS[1:6])


def _editor_rows() -> list[dict[str, object]]:
    return [
        dict(
            zip(
                DECOMPOSITION_HEADERS,
                [
                    90,
                    f"项目{index}",
                    f"M{index}",
                    f"L{index}",
                    f"OP0{index}0",
                    f"工位{index}",
                    f"动作{index}",
                    f"翻译{index}",
                ],
            )
        )
        for index in range(1, 4)
    ]


def test_xlsx_review_upload_roundtrip_preserves_an_added_row_and_renumbers():
    stage = _decomposition_stage(
        [[88, "项目A", "M1", "L1", "OP010", "装配", "原动作"]]
    )
    workbook = load_workbook(BytesIO(stage.decomposition_bytes), data_only=False)
    worksheet = workbook[INPUT_SHEET_NAME]
    worksheet.append(
        [999, "新增项目", "M2", "L2", "OP020", "新增工位", "新增动作", "新增翻译"]
    )
    edited_upload = BytesIO()
    workbook.save(edited_upload)

    uploaded_rows = parse_decomposition_review_upload(
        edited_upload.getvalue(),
        "线下审核版.XLSX",
    )
    assert [row["序号"] for row in uploaded_rows] == [1, 999]
    assert uploaded_rows[1] == dict(
        zip(
            DECOMPOSITION_HEADERS,
            [999, "新增项目", "M2", "L2", "OP020", "新增工位", "新增动作", "新增翻译"],
        )
    )

    reviewed = review_decomposition_rows(stage, uploaded_rows)
    assert [row["序号"] for row in reviewed.decomposition_rows()] == [1, 2]
    assert [row["作业描述"] for row in reviewed.decomposition_rows()] == [
        "原动作",
        "新增动作",
    ]
    assert parse_decomposition_review_upload(
        reviewed.decomposition_bytes,
        reviewed.decomposition_filename,
    ) == reviewed.decomposition_rows()


def test_csv_review_upload_roundtrip_restores_formula_safe_prefixes():
    stage = _decomposition_stage(
        [[1, "项目A", "M1", "L1", "OP010", "装配", "原动作"]]
    )
    edited_row = dict(stage.decomposition_rows()[0])
    edited_row.update(
        {
            "项目名称": "+项目A",
            "产品型号": "-M1",
            "产线": "@L1",
            "作业描述": "=1+1",
            "翻译后作业描述": "=人工确认文本",
        }
    )
    reviewed = review_decomposition_rows(stage, [edited_row])

    downloaded_rows = _csv_rows(reviewed.decomposition_csv_bytes)
    assert downloaded_rows[1][1:4] == ["'+项目A", "'-M1", "'@L1"]
    assert downloaded_rows[1][6:8] == ["'=1+1", "'=人工确认文本"]

    uploaded_rows = parse_decomposition_review_upload(
        reviewed.decomposition_csv_bytes,
        reviewed.decomposition_csv_filename,
    )
    assert uploaded_rows[0]["项目名称"] == "+项目A"
    assert uploaded_rows[0]["产品型号"] == "-M1"
    assert uploaded_rows[0]["产线"] == "@L1"
    assert uploaded_rows[0]["作业描述"] == "=1+1"
    assert uploaded_rows[0]["翻译后作业描述"] == "=人工确认文本"

    roundtripped = review_decomposition_rows(stage, uploaded_rows)
    assert roundtripped.decomposition_rows()[0] == {
        **edited_row,
        "序号": 1,
    }
    assert parse_decomposition_review_upload(
        roundtripped.decomposition_csv_bytes,
        roundtripped.decomposition_csv_filename,
    ) == uploaded_rows


@pytest.mark.parametrize("filename", ["审核版.xls", "审核版.txt", "审核版"])
def test_review_upload_rejects_unsupported_extensions(filename: str):
    with pytest.raises(ExcelInputError, match=r"\.xlsx.*\.csv"):
        parse_decomposition_review_upload(b"not empty", filename)


@pytest.mark.parametrize("suffix", [".xlsx", ".csv"])
def test_review_upload_rejects_a_missing_required_column(suffix: str):
    headers = DECOMPOSITION_HEADERS[:-1]
    if suffix == ".xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = INPUT_SHEET_NAME
        worksheet.append(list(headers))
        worksheet.append([1, "项目A", "M1", "L1", "OP010", "装配", "动作"])
        output = BytesIO()
        workbook.save(output)
        payload = output.getvalue()
    else:
        output = StringIO(newline="")
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerow([1, "项目A", "M1", "L1", "OP010", "装配", "动作"])
        payload = ("\ufeff" + output.getvalue()).encode("utf-8")

    with pytest.raises(ExcelInputError, match="第 1 行必须依次为"):
        parse_decomposition_review_upload(payload, f"缺列{suffix}")


@pytest.mark.parametrize("filename", ["空审核版.xlsx", "空审核版.csv"])
def test_review_upload_rejects_an_empty_file(filename: str):
    with pytest.raises(ExcelInputError, match="为空"):
        parse_decomposition_review_upload(b"", filename)


def test_xlsx_review_upload_requires_the_data_sheet():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "审核表"
    worksheet.append(list(DECOMPOSITION_HEADERS))
    worksheet.append([1, "项目A", "M1", "L1", "OP010", "装配", "动作", "翻译"])
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ExcelInputError, match="缺少.*数据表.*工作表"):
        parse_decomposition_review_upload(output.getvalue(), "审核版.xlsx")


@pytest.mark.parametrize(
    ("cell", "expected_message"),
    [("I1", "表头只能包含 A:H 八列"), ("I2", "第 2 行包含 A:H 之外的数据")],
)
def test_xlsx_review_upload_rejects_a_nonempty_ninth_column(
    cell: str,
    expected_message: str,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = INPUT_SHEET_NAME
    worksheet.append(list(DECOMPOSITION_HEADERS))
    worksheet.append([1, "项目A", "M1", "L1", "OP010", "装配", "动作", "翻译"])
    worksheet[cell] = "第九列数据"
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ExcelInputError, match=expected_message):
        parse_decomposition_review_upload(output.getvalue(), "审核版.xlsx")


@pytest.mark.parametrize(
    "values",
    [
        [1, "项目A", "M1", "L1", "OP010", "装配", "动作", "翻译", "第九列"],
        [1, "项目A", "M1", "L1", "OP010", "装配", "动作"],
    ],
    ids=["ninth-column", "nonempty-short-row"],
)
def test_csv_review_upload_rejects_non_eight_column_data_rows(
    values: list[object],
):
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(DECOMPOSITION_HEADERS)
    writer.writerow(values)
    payload = ("\ufeff" + output.getvalue()).encode("utf-8")

    with pytest.raises(ExcelInputError, match="第 2 行必须正好包含八列"):
        parse_decomposition_review_upload(payload, "审核版.csv")


def test_xlsx_review_upload_rejects_a_real_formula_cell():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = INPUT_SHEET_NAME
    worksheet.append(list(DECOMPOSITION_HEADERS))
    worksheet.append(
        [1, "项目A", "M1", "L1", "OP010", "装配", "=1+1", "人工确认文本"]
    )
    assert worksheet["G2"].data_type == "f"
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ExcelInputError, match="第 2 行第 G 列不能使用公式"):
        parse_decomposition_review_upload(output.getvalue(), "审核版.xlsx")


def test_system_exported_formula_like_xlsx_text_is_accepted():
    stage = _decomposition_stage(
        [[1, "项目A", "M1", "L1", "OP010", "装配", "原动作"]]
    )
    edited_row = dict(stage.decomposition_rows()[0])
    edited_row["作业描述"] = "=1+1"
    edited_row["翻译后作业描述"] = "=人工确认文本"
    reviewed = review_decomposition_rows(stage, [edited_row])

    workbook = load_workbook(
        BytesIO(reviewed.decomposition_bytes),
        data_only=False,
    )
    worksheet = workbook[INPUT_SHEET_NAME]
    assert worksheet["G2"].data_type == "s"
    assert worksheet["H2"].data_type == "s"

    assert parse_decomposition_review_upload(
        reviewed.decomposition_bytes,
        reviewed.decomposition_filename,
    ) == [{**edited_row, "序号": 1}]


@pytest.mark.parametrize("suffix", [".xlsx", ".csv"])
def test_review_upload_rejects_structurally_valid_but_blank_data(suffix: str):
    if suffix == ".xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = INPUT_SHEET_NAME
        worksheet.append(list(DECOMPOSITION_HEADERS))
        worksheet.append([None] * len(DECOMPOSITION_HEADERS))
        output = BytesIO()
        workbook.save(output)
        payload = output.getvalue()
    else:
        output = StringIO(newline="")
        writer = csv.writer(output)
        writer.writerow(DECOMPOSITION_HEADERS)
        writer.writerow([""] * len(DECOMPOSITION_HEADERS))
        payload = ("\ufeff" + output.getvalue()).encode("utf-8")

    with pytest.raises(ExcelInputError, match="没有可解析的动作"):
        parse_decomposition_review_upload(payload, f"全空数据{suffix}")


@pytest.mark.parametrize(
    ("position", "copied_from"),
    [(1, 0), (2, 0), (3, 1), (4, 2)],
)
def test_insert_review_editor_row_copies_adjacent_metadata_and_renumbers(
    position: int,
    copied_from: int,
):
    original = _editor_rows()

    inserted = insert_review_editor_row(original, position)

    assert [row["序号"] for row in inserted] == [1, 2, 3, 4]
    new_row = inserted[position - 1]
    assert _metadata(new_row) == _metadata(original[copied_from])
    assert new_row["作业描述"] == ""
    assert new_row["翻译后作业描述"] == ""
    assert [row["作业描述"] for row in inserted if row is not new_row] == [
        "动作1",
        "动作2",
        "动作3",
    ]


def test_delete_review_editor_rows_removes_selected_rows_and_renumbers():
    rows = _editor_rows() + [
        dict(
            zip(
                DECOMPOSITION_HEADERS,
                [400, "项目4", "M4", "L4", "OP040", "工位4", "动作4", "翻译4"],
            )
        )
    ]

    remaining = delete_review_editor_rows(rows, [4, 2])

    assert [row["序号"] for row in remaining] == [1, 2]
    assert [row["作业描述"] for row in remaining] == ["动作1", "动作3"]
    assert remaining == normalize_review_editor_rows([rows[0], rows[2]])
