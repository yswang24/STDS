"""STDS Excel 批量输入/输出：保留 PF 元数据并生成拆解与工时文件。"""
from __future__ import annotations

import asyncio
import csv
import inspect
import json
import logging
import time
import unicodedata
import zipfile
from dataclasses import dataclass, replace
from io import BytesIO, StringIO
from pathlib import Path
from typing import Awaitable, Callable, Mapping, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from stds.cascade import rules
from stds.cascade.resolver import (
    Deps,
    PartWeightGroupResolution,
    resolve,
    resolve_part_weight_groups,
)
from stds.cascade.rules import normalize
from stds.config.settings import settings
from stds.domain.models import Source, StdsElement, StdsResult
from stds.llm.decompose import decompose_operation
from stds.llm.translate_operation import (
    translate_operation_for_display,
    translate_operation_for_output,
)
from stds.pipeline.operation_analysis import (
    Decomposer,
    OperationSplit,
    Resolver,
    classify_operation_actor,
    resolve_with_actor,
    split_operation,
)
from stds.pipeline.output_schema import (
    CHARTCODE_HEADER,
    CV_HEADER,
    DECISION_HEADER,
    DECISION_REASON_HEADER,
    DECOMPOSITION_HEADERS,
    FREQ_HEADER,
    INPUT_HEADERS,
    JES_HEADER,
    LINE_HEADER,
    NUMBER_HEADER,
    OUTPUT_HEADERS,
    OUTPUT_OPERATION_HEADER,
    PRODUCT_MODEL_HEADER,
    PROJECT_HEADER,
    SOS_HEADER,
    STATION_DESCRIPTION_HEADER,
    STATION_HEADER,
    STDS_HEADER,
    TIME_HEADER,
    TRACE_HEADER,
    TRANSLATED_OPERATION_HEADER,
)
from stds.pipeline.repeated_action import (
    RepeatedActionGroup,
    build_repeated_action_groups,
)
from stds.pipeline.trace_output import (
    EXCEL_CELL_TEXT_LIMIT,
    result_trace_items,
    serialize_trace,
)

logger = logging.getLogger("stds.excel_batch")

INPUT_SHEET_NAME = "数据表"
INPUT_HEADER_ALIASES = {
    NUMBER_HEADER: NUMBER_HEADER,
    "number": NUMBER_HEADER,
    PROJECT_HEADER: PROJECT_HEADER,
    "project_name": PROJECT_HEADER,
    PRODUCT_MODEL_HEADER: PRODUCT_MODEL_HEADER,
    "product_model": PRODUCT_MODEL_HEADER,
    LINE_HEADER: LINE_HEADER,
    "line_name": LINE_HEADER,
    STATION_HEADER: STATION_HEADER,
    "station_op": STATION_HEADER,
    STATION_DESCRIPTION_HEADER: STATION_DESCRIPTION_HEADER,
    "station_description": STATION_DESCRIPTION_HEADER,
    OUTPUT_OPERATION_HEADER: OUTPUT_OPERATION_HEADER,
    "operation": OUTPUT_OPERATION_HEADER,
    "operation_des": OUTPUT_OPERATION_HEADER,
}
PHASE_DECOMPOSE = "拆解"
PHASE_ANALYZE = "工时分析"
PENDING_REVIEW_ACTOR = "待判定"
MAX_REVIEW_UPLOAD_BYTES = 20 * 1024 * 1024
MAX_REVIEW_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_REVIEW_XLSX_ENTRIES = 2000
MAX_REVIEW_ROWS = 100_000


def _effective_actor(operation: str, inherited_actor: str) -> str:
    """明确设备子动作覆盖父工序主体，其余动作继承父级判定。"""
    try:
        if rules.is_explicit_machine_action(operation):
            return "设备"
    except Exception:
        logger.debug(
            "Explicit machine child actor rule failed: operation=%r",
            operation,
            exc_info=True,
        )
    return inherited_actor


def _effective_row_actor(
    operations: Sequence[str],
    inherited_actor: str,
) -> str:
    actors = {
        _effective_actor(operation, inherited_actor)
        for operation in operations
    }
    if not actors:
        return inherited_actor
    if len(actors) == 1:
        return next(iter(actors))
    return "混合"


class ExcelInputError(ValueError):
    """上传的工作簿缺少可解析结构。"""


@dataclass(frozen=True)
class ExcelInputRow:
    sheet_name: str
    row_index: int
    operation: str
    number: object
    project_name: object
    product_model: object
    line_name: object
    station_op: object
    station_description: object
    norm_key: str


@dataclass
class ExcelDetailResult:
    input_row: ExcelInputRow
    split: OperationSplit
    child_index: int
    operation: str
    output_number: int
    display_operation: Optional[str] = None
    result: Optional[StdsResult] = None
    error: Optional[str] = None
    repeated_action_trace: Optional[tuple[str, str, str]] = None

    @property
    def effective_actor(self) -> str:
        return _effective_actor(self.operation, self.split.actor)

    @property
    def status(self) -> str:
        if self.error is not None or self.result is None:
            return "失败"
        if (
            self.split.needs_review
            or self.result.needs_review
            or self.result.source == Source.UNRESOLVED
        ):
            return "待复核"
        return "成功"

    @property
    def child_count(self) -> int:
        return len(self.split.operations)

    def as_decomposition_preview(self) -> dict:
        """返回 PF 拆解原文以及用于最终输出的翻译结果。"""
        return dict(zip(DECOMPOSITION_HEADERS, self.decomposition_values()))

    def decomposition_values(self) -> list:
        """拆解文件保留原文，并新增最终输出所使用的翻译后描述。"""
        return [
            self.output_number,
            self.input_row.project_name,
            self.input_row.product_model,
            self.input_row.line_name,
            self.input_row.station_op,
            self.input_row.station_description,
            self.operation,
            self.output_operation,
        ]

    def as_preview(self) -> dict:
        return {
            **self.output_row(),
            "主体类型": self.effective_actor,
            TRACE_HEADER: _detail_trace(self),
            "状态": self.status,
        }

    def analysis_values(self) -> tuple[object, object, object, object, object]:
        """返回五个分析字段；无人工工时分析结果时统一输出 NA。"""
        result = self.result
        if (
            self.effective_actor == "设备"
            or result is None
            or result.source in {Source.MACHINE, Source.UNRESOLVED}
        ):
            return ("NA", "NA", "NA", "NA", "NA")
        time_value = self.time_value()
        return (
            result.decision or "NA",
            result.chartcode or "NA",
            result.cv or "NA",
            result.freq if result.freq is not None else "NA",
            time_value if time_value is not None else "NA",
        )

    def output_values(self) -> list:
        """返回工时生成模板 A:O 的十五列值。"""
        decision, chartcode, cv, freq, time_value = self.analysis_values()
        return [
            self.output_number,
            self.input_row.project_name,
            self.input_row.product_model,
            self.input_row.line_name,
            self.input_row.station_op,
            self.input_row.station_description,
            "NA",
            "NA",
            self.output_operation,
            decision,
            chartcode,
            cv,
            freq,
            time_value,
            _detail_decision_reason(self),
        ]

    @property
    def output_operation(self) -> str:
        """最终展示文本；分析、检索和 trace 仍保留拆解后的原文。"""
        return self.display_operation or self.operation

    def output_row(self) -> dict:
        """前端明细与下载工作簿共用的十五列记录。"""
        return dict(zip(OUTPUT_HEADERS, self.output_values()))

    def time_value(self) -> Optional[float]:
        if (
            self.effective_actor == "设备"
            or self.split.needs_review
            or self.result is None
        ):
            return None
        return _result_time(self.result)


@dataclass
class ExcelRowResult:
    input_row: ExcelInputRow
    split: OperationSplit
    details: list[ExcelDetailResult]

    @property
    def effective_actor(self) -> str:
        actors = {detail.effective_actor for detail in self.details}
        if not actors:
            return self.split.actor
        if len(actors) == 1:
            return next(iter(actors))
        return "混合"

    @property
    def status(self) -> str:
        statuses = {detail.status for detail in self.details}
        if "失败" in statuses:
            return "失败"
        if "待复核" in statuses:
            return "待复核"
        return "成功"

    def decision_value(self) -> str:
        if len(self.details) == 1:
            result = self.details[0].result
            return result.decision if result else ""
        values = [
            {
                "拆解序号": f"{detail.child_index}/{detail.child_count}",
                "operation": detail.operation,
                DECISION_HEADER: detail.result.decision if detail.result else "",
            }
            for detail in self.details
        ]
        return _fit_excel_text(json.dumps(values, ensure_ascii=False, separators=(",", ":")))

    def trace_value(self) -> str:
        trace: list = [
            (
                "拆解",
                json.dumps(self.split.operations, ensure_ascii=False),
                f"主体={self.effective_actor}; 来源={self.split.source}",
            )
        ]
        if self.split.error:
            trace.append(("拆解待复核", "回退为原动作", self.split.error))
        for detail in self.details:
            prefix = f"{detail.child_index}/{detail.child_count}"
            trace.extend(_detail_result_trace_items(detail, prefix=prefix))
        return serialize_trace(trace)

    def time_value(self) -> Optional[float]:
        non_device_details = [
            detail
            for detail in self.details
            if detail.effective_actor != "设备"
        ]
        if not non_device_details:
            return None
        values = [detail.time_value() for detail in non_device_details]
        if any(value is None for value in values):
            return None
        return round(sum(value for value in values if value is not None), 2)

    def as_preview(self) -> dict:
        return {
            "工作表": self.input_row.sheet_name,
            "Excel行": self.input_row.row_index,
            "operation": self.input_row.operation,
            "主体类型": self.effective_actor,
            "拆解数量": len(self.details),
            DECISION_HEADER: self.decision_value(),
            TRACE_HEADER: self.trace_value(),
            TIME_HEADER: self.time_value(),
            "状态": self.status,
        }


@dataclass
class ExcelDecompositionOutput:
    """拆解和翻译阶段的可审核快照；此时尚未执行任何工时分析。"""

    source_bytes: bytes
    source_name: str
    decomposition_bytes: bytes
    decomposition_filename: str
    rows: list[ExcelRowResult]
    timings: list["ExcelProgress"]
    total_elapsed_s: float
    decompose_elapsed_s: float

    def decomposition_rows(self) -> list[dict]:
        return [
            detail.as_decomposition_preview()
            for row in self.rows
            for detail in row.details
        ]

    def decomposition_display_rows(self) -> list[dict]:
        return _display_records(
            DECOMPOSITION_HEADERS,
            _decomposition_records(self.rows),
        )

    @property
    def decomposition_csv_bytes(self) -> bytes:
        return _write_csv(DECOMPOSITION_HEADERS, _decomposition_records(self.rows))

    @property
    def decomposition_csv_filename(self) -> str:
        return build_csv_filename(self.decomposition_filename)

    @property
    def detail_count(self) -> int:
        return sum(len(row.details) for row in self.rows)

    @property
    def total_count(self) -> int:
        return len(self.rows)

    def timing_rows(self) -> list[dict]:
        return [timing.as_preview() for timing in self.timings]


@dataclass
class ExcelBatchOutput:
    output_bytes: bytes
    output_filename: str
    decomposition_bytes: bytes
    decomposition_filename: str
    rows: list[ExcelRowResult]
    timings: list["ExcelProgress"]
    total_elapsed_s: float
    decompose_elapsed_s: float
    analysis_elapsed_s: float
    detail_sheet_name: str

    @property
    def processed_count(self) -> int:
        return sum(any(detail.result is not None for detail in row.details) for row in self.rows)

    @property
    def failed_count(self) -> int:
        return sum(row.status == "失败" for row in self.rows)

    @property
    def review_count(self) -> int:
        return sum(row.status == "待复核" for row in self.rows)

    def preview_rows(self) -> list[dict]:
        return [row.as_preview() for row in self.rows]

    def decomposition_rows(self) -> list[dict]:
        return [
            detail.as_decomposition_preview()
            for row in self.rows
            for detail in row.details
        ]

    def decomposition_display_rows(self) -> list[dict]:
        return _display_records(
            DECOMPOSITION_HEADERS,
            _decomposition_records(self.rows),
        )

    @property
    def decomposition_csv_bytes(self) -> bytes:
        return _write_csv(DECOMPOSITION_HEADERS, _decomposition_records(self.rows))

    @property
    def decomposition_csv_filename(self) -> str:
        return build_csv_filename(self.decomposition_filename)

    def detail_preview_rows(self) -> list[dict]:
        return [detail.output_row() for row in self.rows for detail in row.details]

    def detail_display_rows(self) -> list[dict]:
        return _display_records(OUTPUT_HEADERS, _result_records(self.rows))

    @property
    def output_csv_bytes(self) -> bytes:
        return _write_csv(OUTPUT_HEADERS, _result_records(self.rows))

    @property
    def output_csv_filename(self) -> str:
        return build_csv_filename(self.output_filename)

    @property
    def detail_count(self) -> int:
        return sum(len(row.details) for row in self.rows)

    @property
    def total_count(self) -> int:
        return len(self.rows)

    @property
    def average_elapsed_s(self) -> float:
        return self.total_elapsed_s / self.total_count if self.total_count else 0.0

    def timing_rows(self) -> list[dict]:
        return [timing.as_preview() for timing in self.timings]


@dataclass(frozen=True)
class ExcelProgress:
    phase: str
    completed_rows: int
    total_rows: int
    operation: str
    affected_rows: int
    item_elapsed_s: float
    total_elapsed_s: float
    generated_operations: tuple[str, ...] = ()
    actor: str = ""
    sheet_name: str = ""
    row_index: Optional[int] = None
    number: object = None
    project_name: object = None
    product_model: object = None
    line_name: object = None
    station_op: object = None
    station_description: object = None
    child_index: Optional[int] = None
    child_count: Optional[int] = None

    @property
    def overall_ratio(self) -> float:
        phase_ratio = self.completed_rows / self.total_rows if self.total_rows else 1.0
        if self.phase == PHASE_DECOMPOSE:
            return min(0.5, phase_ratio * 0.5)
        return min(1.0, 0.5 + phase_ratio * 0.5)

    def as_preview(self) -> dict:
        return {
            "阶段": self.phase,
            "完成进度": f"{self.completed_rows}/{self.total_rows}",
            "工作表": self.sheet_name,
            "Excel行": self.row_index,
            OUTPUT_OPERATION_HEADER: self.operation,
            "主体类型": self.actor,
            "拆解序号": (
                f"{self.child_index}/{self.child_count}"
                if self.child_index is not None and self.child_count is not None
                else ""
            ),
            "本条耗时（秒）": round(self.item_elapsed_s, 2),
            "覆盖 Excel 行数": self.affected_rows,
            "累计耗时（秒）": round(self.total_elapsed_s, 2),
        }


ProgressCallback = Callable[[ExcelProgress], object]
OutputTranslator = Callable[[str], Awaitable[str]]


def _clean_header(value: object) -> str:
    """清理表头中的 BOM、零宽字符及各种空白，不做字段名称映射。"""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    invisible = {"\ufeff", "\u200b", "\u200c", "\u200d", "\u2060"}
    return "".join(char for char in text if not char.isspace() and char not in invisible)


def _canonical_header(value: object) -> str:
    cleaned = _clean_header(value)
    return INPUT_HEADER_ALIASES.get(cleaned.casefold(), cleaned)


def _load_inputs(excel_bytes: bytes):
    """读取 PF 清单 A:G，并以“作业描述”作为待拆解文本。"""
    if not excel_bytes:
        raise ExcelInputError("上传的 Excel 文件为空")
    try:
        workbook = load_workbook(BytesIO(excel_bytes), data_only=False)
        values_workbook = load_workbook(BytesIO(excel_bytes), data_only=True)
    except Exception as exc:
        raise ExcelInputError("无法读取该文件，请确认它是有效的 .xlsx 工作簿") from exc

    if INPUT_SHEET_NAME not in workbook.sheetnames:
        raise ExcelInputError(f"固定模板缺少“{INPUT_SHEET_NAME}”工作表")

    ws = workbook[INPUT_SHEET_NAME]
    values_ws = values_workbook[INPUT_SHEET_NAME]
    actual_headers = tuple(
        _canonical_header(ws.cell(1, col).value)
        for col in range(1, len(INPUT_HEADERS) + 1)
    )
    if actual_headers != INPUT_HEADERS:
        raise ExcelInputError(
            "数据表第 1 行必须依次为："
            "序号、项目名称、产品型号、产线、工位号、工位描述、作业描述"
        )

    input_rows: list[ExcelInputRow] = []
    has_formula_operations = False
    operation_col = INPUT_HEADERS.index(OUTPUT_OPERATION_HEADER) + 1

    for row_index in range(2, ws.max_row + 1):
        operation_cell = ws.cell(row_index, operation_col)
        if operation_cell.data_type == "f":
            has_formula_operations = True
            raw_operation = values_ws.cell(row_index, operation_col).value
            if raw_operation is None:
                raise ExcelInputError(
                    f"数据表第 {row_index} 行的作业描述是公式，"
                    "但文件中没有已计算值；请先在 Excel 中重新计算并保存"
                )
        else:
            raw_operation = operation_cell.value
        if raw_operation is None or not str(raw_operation).strip():
            continue

        number_value = ws.cell(row_index, 1).value
        project_value = ws.cell(row_index, 2).value
        product_model_value = ws.cell(row_index, 3).value
        line_value = ws.cell(row_index, 4).value
        station_value = ws.cell(row_index, 5).value
        station_description_value = ws.cell(row_index, 6).value
        if number_value in (None, ""):
            raise ExcelInputError(f"数据表第 {row_index} 行的序号不能为空")
        if station_value in (None, ""):
            raise ExcelInputError(f"数据表第 {row_index} 行的工位号不能为空")

        operation = str(raw_operation).strip()
        input_rows.append(
            ExcelInputRow(
                sheet_name=INPUT_SHEET_NAME,
                row_index=row_index,
                operation=operation,
                number=number_value,
                project_name=project_value,
                product_model=product_model_value,
                line_name=line_value,
                station_op=station_value,
                station_description=station_description_value,
                norm_key=normalize(operation) or operation,
            )
        )

    if not input_rows:
        raise ExcelInputError("作业描述字段下没有可解析的输入")
    if has_formula_operations:
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
    return workbook, input_rows


def _fit_excel_text(value: str) -> str:
    if len(value) <= EXCEL_CELL_TEXT_LIMIT:
        return value
    return value[: EXCEL_CELL_TEXT_LIMIT - 1] + "…"


def _detail_trace(detail: ExcelDetailResult) -> str:
    trace = [
        (
            "拆解",
            detail.operation,
            f"{detail.child_index}/{detail.child_count}; 主体={detail.effective_actor}; 来源={detail.split.source}",
        )
    ]
    if detail.split.error:
        trace.append(("拆解待复核", "回退为原动作", detail.split.error))
    trace.extend(_detail_result_trace_items(detail))
    return serialize_trace(trace)


def _detail_result_trace_items(
    detail: ExcelDetailResult,
    *,
    prefix: str = "",
) -> list:
    """返回明细分析 trace；resolver 抛错时仍保留重复动作的一致性来源。"""
    items = result_trace_items(detail.result, detail.error, prefix=prefix)
    repeated_trace = detail.repeated_action_trace
    if repeated_trace is None:
        return items
    result_has_trace = detail.result is not None and repeated_trace in (
        detail.result.trace or []
    )
    if result_has_trace:
        return items
    variable, choice, reason = repeated_trace
    if prefix:
        variable = f"{prefix}:{variable}"
    return [(variable, choice, reason), *items]


def _detail_decision_reason(detail: ExcelDetailResult) -> str:
    return serialize_trace(_detail_result_trace_items(detail))


def _result_time(result: StdsResult) -> Optional[float]:
    if result.needs_review or result.source == Source.UNRESOLVED:
        return None
    return result.time_s


async def _translate_output_operations(
    rows: list[ExcelRowResult],
    translator: OutputTranslator,
    sem: asyncio.Semaphore,
) -> None:
    """在最终展示前按原文去重翻译；失败只回退展示原文，不影响分析结果。"""
    grouped_details: dict[str, list[ExcelDetailResult]] = {}
    for row in rows:
        for detail in row.details:
            grouped_details.setdefault(detail.operation, []).append(detail)

    async def translate_group(
        operation: str,
        details: list[ExcelDetailResult],
    ) -> None:
        async with sem:
            translated = await translate_operation_for_display(
                operation,
                translator=translator,
            )
        for detail in details:
            detail.display_operation = translated

    await asyncio.gather(
        *(
            translate_group(operation, details)
            for operation, details in grouped_details.items()
        )
    )


def _reset_output_sheet(workbook):
    """清空数据表的旧数据与结构，避免尾随空列或旧表格残留。"""
    ws = workbook[INPUT_SHEET_NAME]
    for table_name in list(ws.tables):
        del ws.tables[table_name]
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    ws.delete_rows(1, max(1, ws.max_row))
    ws.delete_cols(1, max(1, ws.max_column))
    ws.column_dimensions.clear()
    ws.row_dimensions.clear()
    ws.freeze_panes = None
    ws.auto_filter.ref = None
    return ws


def _write_records(
    workbook,
    headers: tuple[str, ...],
    records: list[list],
    *,
    widths: Optional[list[Optional[float]]] = None,
    numeric_formats: Optional[dict[str, str]] = None,
) -> tuple[bytes, str]:
    """按样例的简洁表格样式写入记录，并保证物理列数等于表头数。"""
    ws = _reset_output_sheet(workbook)
    header_font = Font(name="Calibri", size=10, bold=False)
    body_font = Font(name="DengXian", size=11, bold=False)
    center_vertical = Alignment(vertical="center")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = header_font
        cell.alignment = center_vertical
    ws.row_dimensions[1].height = 25.5

    for output_row, values in enumerate(records, start=2):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(output_row, col, value)
            if isinstance(value, str) and value.startswith("="):
                # 审核表中的用户文本必须保持为文本，不能在下载后变成公式。
                cell.data_type = "s"
            cell.font = body_font
            cell.alignment = center_vertical
        ws.row_dimensions[output_row].height = 25.5

    for col, width in enumerate(widths or [], start=1):
        if width is not None:
            ws.column_dimensions[get_column_letter(col)].width = width

    for header, number_format in (numeric_formats or {}).items():
        col = headers.index(header) + 1
        for row_index in range(2, len(records) + 2):
            if isinstance(ws.cell(row_index, col).value, (int, float)):
                ws.cell(row_index, col).number_format = number_format

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), INPUT_SHEET_NAME


def _decomposition_records(rows: list[ExcelRowResult]) -> list[list]:
    return [
        detail.decomposition_values()
        for row in rows
        for detail in row.details
    ]


def _result_records(rows: list[ExcelRowResult]) -> list[list]:
    return [detail.output_values() for row in rows for detail in row.details]


def _export_text(header: str, value: object) -> str:
    if value is None:
        return ""
    if header == TIME_HEADER and isinstance(value, (int, float)):
        return f"{value:.2f}"
    if header == FREQ_HEADER and isinstance(value, (int, float)):
        return f"{value:.2f}".rstrip("0").rstrip(".")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _display_records(headers: tuple[str, ...], records: list[list]) -> list[dict]:
    """生成与 CSV 相同列序和值文本的前端展示记录。"""
    return [
        {
            header: _export_text(header, value)
            for header, value in zip(headers, values)
        }
        for values in records
    ]


def _write_csv(headers: tuple[str, ...], records: list[list]) -> bytes:
    """以 UTF-8 BOM 输出 CSV；数据与对应 XLSX 共用同一 records。"""
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\r\n")
    writer.writerow(headers)
    for values in records:
        csv_values = []
        for header, value in zip(headers, values):
            text = _export_text(header, value)
            # 防止用户输入的文本在 Excel 打开 CSV 时被当作公式执行。
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                text = "'" + text
            csv_values.append(text)
        writer.writerow(csv_values)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _write_decomposition(
    workbook,
    rows: list[ExcelRowResult],
) -> tuple[bytes, str]:
    """生成八列 PF 拆解文件：七列源结构加翻译后作业描述。"""
    records = _decomposition_records(rows)
    # 原文和翻译列同时存在时适当加宽，避免两列文字相互遮挡。
    return _write_records(
        workbook,
        DECOMPOSITION_HEADERS,
        records,
        widths=[19.0, None, None, None, None, None, 45.0, 45.0],
    )


def _write_results(workbook, rows: list[ExcelRowResult]) -> tuple[bytes, str]:
    """生成 3.STDS-工时生成.xlsx 的 A:O，原模板后三列不创建。"""
    records = _result_records(rows)
    return _write_records(
        workbook,
        OUTPUT_HEADERS,
        records,
        widths=[19.0] * (len(OUTPUT_HEADERS) - 1) + [80.0],
        numeric_formats={FREQ_HEADER: "0.##", TIME_HEADER: "0.00"},
    )


def _build_stage_filename(
    source_name: str,
    *,
    numbered_prefix: str,
    fallback_suffix: str,
) -> str:
    safe_name = Path(source_name or "input.xlsx").name
    stem = Path(safe_name).stem or "input"
    source_prefix = "1.STDS-PF清单"
    if stem.casefold().startswith(source_prefix.casefold()):
        stem = numbered_prefix + stem[len(source_prefix) :]
        return f"{stem}.xlsx"
    return f"{stem}_{fallback_suffix}.xlsx"


def build_decomposition_filename(source_name: str) -> str:
    return _build_stage_filename(
        source_name,
        numbered_prefix="2.STDS-PF拆解",
        fallback_suffix="PF拆解",
    )


def build_output_filename(source_name: str) -> str:
    return _build_stage_filename(
        source_name,
        numbered_prefix="3.STDS-工时生成",
        fallback_suffix="工时生成",
    )


def build_csv_filename(xlsx_filename: str) -> str:
    return f"{Path(xlsx_filename).stem}.csv"


async def _notify_excel_progress(
    on_progress: Optional[ProgressCallback],
    progress: ExcelProgress,
) -> None:
    if on_progress is None:
        return
    try:
        maybe_awaitable = on_progress(progress)
        if inspect.isawaitable(maybe_awaitable):
            await maybe_awaitable
    except Exception:
        logger.debug("Excel progress callback failed", exc_info=True)


async def decompose_excel_bytes(
    excel_bytes: bytes,
    source_name: str,
    deps: Deps,
    *,
    decomposer: Decomposer = decompose_operation,
    translator: OutputTranslator = translate_operation_for_output,
    concurrency: Optional[int] = None,
    on_progress: Optional[ProgressCallback] = None,
) -> ExcelDecompositionOutput:
    """只执行拆解与展示文本翻译，供人工审核；不会调用工时 resolver。"""
    batch_started = time.perf_counter()
    workbook, input_rows = _load_inputs(excel_bytes)
    timings: list[ExcelProgress] = []
    sem = asyncio.Semaphore(max(1, concurrency or settings.CONCURRENCY_LIMIT))

    grouped_rows: dict[str, list[ExcelInputRow]] = {}
    for input_row in input_rows:
        grouped_rows.setdefault(input_row.norm_key, []).append(input_row)

    decompose_completed = 0

    async def split_group(
        norm_key: str,
        group: list[ExcelInputRow],
    ) -> tuple[str, OperationSplit]:
        nonlocal decompose_completed
        representative = group[0]
        item_started: Optional[float] = None
        split = OperationSplit(
            actor="人工",
            operations=(representative.operation,),
            source="拆解失败回退",
            needs_review=True,
        )
        try:
            async with sem:
                item_started = time.perf_counter()
                split = await split_operation(
                    representative.operation,
                    deps,
                    decomposer=decomposer,
                )
                return norm_key, split
        except Exception as exc:
            logger.exception(
                "Excel operation decomposition failed: operation=%r rows=%s",
                representative.operation,
                [(input_row.sheet_name, input_row.row_index) for input_row in group],
            )
            split = replace(split, error=f"{type(exc).__name__}: {exc}")
            return norm_key, split
        finally:
            item_elapsed_s = (
                time.perf_counter() - item_started if item_started is not None else 0.0
            )
            per_row_elapsed_s = item_elapsed_s / len(group)
            for input_row in group:
                decompose_completed += 1
                progress = ExcelProgress(
                    phase=PHASE_DECOMPOSE,
                    completed_rows=decompose_completed,
                    total_rows=len(input_rows),
                    operation=input_row.operation,
                    affected_rows=1,
                    item_elapsed_s=per_row_elapsed_s,
                    total_elapsed_s=time.perf_counter() - batch_started,
                    generated_operations=split.operations,
                    actor=_effective_row_actor(split.operations, split.actor),
                    sheet_name=input_row.sheet_name,
                    row_index=input_row.row_index,
                    number=input_row.number,
                    project_name=input_row.project_name,
                    product_model=input_row.product_model,
                    line_name=input_row.line_name,
                    station_op=input_row.station_op,
                    station_description=input_row.station_description,
                )
                timings.append(progress)
                await _notify_excel_progress(on_progress, progress)

    split_pairs = await asyncio.gather(
        *(split_group(norm_key, group) for norm_key, group in grouped_rows.items())
    )
    split_by_key = dict(split_pairs)

    rows: list[ExcelRowResult] = []
    output_number = 0
    for input_row in input_rows:
        split = split_by_key[input_row.norm_key]
        details = []
        for index, operation in enumerate(split.operations, start=1):
            output_number += 1
            details.append(
                ExcelDetailResult(
                    input_row=input_row,
                    split=split,
                    child_index=index,
                    operation=operation,
                    output_number=output_number,
                )
            )
        rows.append(ExcelRowResult(input_row=input_row, split=split, details=details))

    # 翻译在人工审核前完成；工时分析始终使用第 G 列拆解原文。
    await _translate_output_operations(rows, translator, sem)
    decomposition_bytes, _ = _write_decomposition(workbook, rows)
    decompose_elapsed_s = time.perf_counter() - batch_started
    return ExcelDecompositionOutput(
        source_bytes=bytes(excel_bytes),
        source_name=source_name,
        decomposition_bytes=decomposition_bytes,
        decomposition_filename=build_decomposition_filename(source_name),
        rows=rows,
        timings=timings,
        total_elapsed_s=decompose_elapsed_s,
        decompose_elapsed_s=decompose_elapsed_s,
    )


def _is_blank_review_value(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    try:
        unequal = value != value
        if bool(unequal):
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip() in {"", "<NA>", "NaT"}


def _clean_review_value(value: object) -> object:
    if _is_blank_review_value(value):
        return None
    return value.strip() if isinstance(value, str) else value


def _validate_review_upload_headers(values: Sequence[object]) -> None:
    actual_headers = tuple(_clean_header(value) for value in values)
    if actual_headers != DECOMPOSITION_HEADERS:
        raise ExcelInputError(
            "人工审核拆解文件第 1 行必须依次为："
            + "、".join(DECOMPOSITION_HEADERS)
        )


def _restore_review_csv_text(value: str) -> str:
    if len(value) >= 2 and value[0] == "'" and value[1] in ("=", "+", "-", "@"):
        return value[1:]
    return value


def parse_decomposition_review_upload(
    file_bytes: bytes,
    filename: str,
) -> list[dict[str, object]]:
    """读取线下修改后的八列审核版 XLSX/CSV，返回 canonical records。"""
    if not file_bytes:
        raise ExcelInputError("上传的人工审核拆解文件为空")
    if len(file_bytes) > MAX_REVIEW_UPLOAD_BYTES:
        raise ExcelInputError("人工审核拆解文件不能超过 20 MB")

    suffix = Path(filename or "").suffix.casefold()
    if suffix == ".xlsx":
        try:
            with zipfile.ZipFile(BytesIO(file_bytes)) as archive:
                entries = archive.infolist()
                if len(entries) > MAX_REVIEW_XLSX_ENTRIES:
                    raise ExcelInputError("人工审核 XLSX 包含过多内部文件")
                if (
                    sum(entry.file_size for entry in entries)
                    > MAX_REVIEW_XLSX_UNCOMPRESSED_BYTES
                ):
                    raise ExcelInputError("人工审核 XLSX 解压后不能超过 100 MB")
            workbook = load_workbook(
                BytesIO(file_bytes),
                data_only=False,
                read_only=False,
            )
        except ExcelInputError:
            raise
        except Exception as exc:
            raise ExcelInputError(
                "无法读取人工审核拆解文件，请确认它是有效的 .xlsx 工作簿"
            ) from exc
        try:
            if INPUT_SHEET_NAME not in workbook.sheetnames:
                raise ExcelInputError(
                    f"人工审核拆解文件缺少“{INPUT_SHEET_NAME}”工作表"
                )
            ws = workbook[INPUT_SHEET_NAME]
            if ws.max_row > MAX_REVIEW_ROWS + 1:
                raise ExcelInputError(
                    f"人工审核拆解文件最多支持 {MAX_REVIEW_ROWS} 条动作"
                )
            _validate_review_upload_headers(
                [
                    ws.cell(1, column).value
                    for column in range(1, len(DECOMPOSITION_HEADERS) + 1)
                ]
            )
            if any(
                not _is_blank_review_value(ws.cell(1, column).value)
                for column in range(
                    len(DECOMPOSITION_HEADERS) + 1,
                    ws.max_column + 1,
                )
            ):
                raise ExcelInputError("人工审核拆解文件表头只能包含 A:H 八列")

            records = []
            for row_index in range(2, ws.max_row + 1):
                cells = [
                    ws.cell(row_index, column)
                    for column in range(1, len(DECOMPOSITION_HEADERS) + 1)
                ]
                extra_values = [
                    ws.cell(row_index, column).value
                    for column in range(
                        len(DECOMPOSITION_HEADERS) + 1,
                        ws.max_column + 1,
                    )
                ]
                if any(
                    not _is_blank_review_value(value)
                    for value in extra_values
                ):
                    raise ExcelInputError(
                        f"人工审核拆解文件第 {row_index} 行包含 A:H 之外的数据"
                    )
                if all(_is_blank_review_value(cell.value) for cell in cells):
                    continue
                for column, cell in enumerate(cells, start=1):
                    if cell.data_type == "f":
                        raise ExcelInputError(
                            f"人工审核拆解文件第 {row_index} 行"
                            f"第 {get_column_letter(column)} 列不能使用公式"
                        )
                records.append(
                    {
                        header: cell.value
                        for header, cell in zip(DECOMPOSITION_HEADERS, cells)
                    }
                )
        finally:
            workbook.close()
    elif suffix == ".csv":
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ExcelInputError(
                "人工审核 CSV 必须使用 UTF-8 编码"
            ) from exc
        try:
            csv_rows = csv.reader(StringIO(text))
            header_row = next(csv_rows)
            _validate_review_upload_headers(header_row)
            records = []
            for row_index, values in enumerate(csv_rows, start=2):
                if not values or all(_is_blank_review_value(value) for value in values):
                    continue
                if len(values) != len(DECOMPOSITION_HEADERS):
                    raise ExcelInputError(
                        f"人工审核 CSV 第 {row_index} 行必须正好包含八列"
                    )
                if len(records) >= MAX_REVIEW_ROWS:
                    raise ExcelInputError(
                        f"人工审核拆解文件最多支持 {MAX_REVIEW_ROWS} 条动作"
                    )
                records.append(
                    {
                        header: _restore_review_csv_text(value)
                        for header, value in zip(DECOMPOSITION_HEADERS, values)
                    }
                )
        except (StopIteration, csv.Error) as exc:
            raise ExcelInputError("无法读取人工审核 CSV 文件") from exc
    else:
        raise ExcelInputError("人工审核拆解文件仅支持 .xlsx 或 .csv")

    if not any(
        not all(
            _is_blank_review_value(record.get(header))
            for header in DECOMPOSITION_HEADERS[1:]
        )
        for record in records
    ):
        raise ExcelInputError("人工审核拆解文件中没有可解析的动作")
    return records


def _review_actor(
    operation: str,
    actor_by_norm_key: Mapping[str, str],
) -> str:
    """按动作键复用主体；歧义的新动作留给异步分析阶段完整判定。"""
    if _effective_actor(operation, "") == "设备":
        return "设备"
    norm_key = normalize(operation) or operation
    if norm_key in actor_by_norm_key:
        return actor_by_norm_key[norm_key]
    try:
        machine = rules.rule_machine(operation)
    except Exception:
        logger.debug("Reviewed operation actor rule failed", exc_info=True)
        machine = None
    if machine is None:
        return PENDING_REVIEW_ACTOR
    return "设备" if machine else "人工"


def review_decomposition_rows(
    decomposition: ExcelDecompositionOutput,
    records: Sequence[Mapping[str, object]],
) -> ExcelDecompositionOutput:
    """校验前端审核记录、按当前顺序重编号，并生成可下载的八列工作簿。"""
    actor_candidates: dict[str, set[str]] = {}
    for row in decomposition.rows:
        for detail in row.details:
            if (
                detail.split.needs_review
                or detail.split.error
                or detail.split.actor not in {"人工", "设备"}
            ):
                continue
            norm_key = normalize(detail.operation) or detail.operation
            actor_candidates.setdefault(norm_key, set()).add(detail.split.actor)
    actor_by_norm_key = {
        norm_key: next(iter(actors))
        for norm_key, actors in actor_candidates.items()
        if len(actors) == 1
    }
    reviewed_rows: list[ExcelRowResult] = []

    for editor_index, record in enumerate(records, start=1):
        if not isinstance(record, Mapping):
            raise ExcelInputError(f"人工审核表第 {editor_index} 行格式无效")
        missing_headers = [
            header for header in DECOMPOSITION_HEADERS[1:] if header not in record
        ]
        if missing_headers:
            raise ExcelInputError(
                "人工审核表缺少字段：" + "、".join(missing_headers)
            )

        editable_values = [record.get(header) for header in DECOMPOSITION_HEADERS[1:]]
        if all(_is_blank_review_value(value) for value in editable_values):
            continue

        cleaned_record = {
            header: _clean_review_value(record.get(header))
            for header in DECOMPOSITION_HEADERS[1:]
        }
        for header, value in cleaned_record.items():
            if isinstance(value, str) and len(value) > EXCEL_CELL_TEXT_LIMIT:
                raise ExcelInputError(
                    f"人工审核表第 {editor_index} 行的{header}超过 Excel 单元格长度上限"
                )

        station_op = cleaned_record[STATION_HEADER]
        if station_op is None:
            raise ExcelInputError(f"人工审核表第 {editor_index} 行的工位号不能为空")
        operation_value = cleaned_record[OUTPUT_OPERATION_HEADER]
        if operation_value is None:
            raise ExcelInputError(f"人工审核表第 {editor_index} 行的作业描述不能为空")
        translated_value = cleaned_record[TRANSLATED_OPERATION_HEADER]
        if translated_value is None:
            raise ExcelInputError(
                f"人工审核表第 {editor_index} 行的翻译后作业描述不能为空"
            )

        operation = str(operation_value)
        translated_operation = str(translated_value)
        output_number = len(reviewed_rows) + 1
        actor = _review_actor(operation, actor_by_norm_key)
        split = OperationSplit(
            actor=actor,
            operations=(operation,),
            source=(
                "人工审核确认（待主体判定）"
                if actor == PENDING_REVIEW_ACTOR
                else "人工审核确认"
            ),
        )
        input_row = ExcelInputRow(
            sheet_name=INPUT_SHEET_NAME,
            row_index=output_number + 1,
            operation=operation,
            number=output_number,
            project_name=cleaned_record[PROJECT_HEADER],
            product_model=cleaned_record[PRODUCT_MODEL_HEADER],
            line_name=cleaned_record[LINE_HEADER],
            station_op=station_op,
            station_description=cleaned_record[STATION_DESCRIPTION_HEADER],
            norm_key=normalize(operation) or operation,
        )
        detail = ExcelDetailResult(
            input_row=input_row,
            split=split,
            child_index=1,
            operation=operation,
            output_number=output_number,
            display_operation=translated_operation,
        )
        reviewed_rows.append(
            ExcelRowResult(input_row=input_row, split=split, details=[detail])
        )

    if not reviewed_rows:
        raise ExcelInputError("人工审核表至少需要保留一条拆解动作")

    # 页面“确认但未修改”时恢复原父子边界，否则每个审核行会退化成独立父工序，
    # 同一父工序下的共享重量上下文就会丢失。发生增删、改写或重排时不猜边界。
    original_details = [
        detail
        for row in decomposition.rows
        for detail in row.details
    ]
    reviewed_details = [
        row.details[0]
        for row in reviewed_rows
    ]
    preserve_parent_boundaries = (
        len(original_details) == len(reviewed_details)
        and all(
            original.decomposition_values()[1:]
            == reviewed.decomposition_values()[1:]
            for original, reviewed in zip(
                original_details,
                reviewed_details,
            )
        )
    )
    if preserve_parent_boundaries:
        regrouped_rows = []
        cursor = 0
        for original_row in decomposition.rows:
            child_count = len(original_row.details)
            child_rows = reviewed_rows[cursor : cursor + child_count]
            cursor += child_count
            child_details = [row.details[0] for row in child_rows]
            actors = {
                detail.split.actor
                for detail in child_details
                if detail.effective_actor != "设备"
            }
            if len(actors) == 1:
                actor = next(iter(actors))
            elif not actors:
                actor = "设备"
            else:
                actor = PENDING_REVIEW_ACTOR
            shared_split = OperationSplit(
                actor=actor,
                operations=tuple(
                    detail.operation for detail in child_details
                ),
                source=child_details[0].split.source,
                needs_review=any(
                    detail.split.needs_review for detail in child_details
                ),
                error=next(
                    (
                        detail.split.error
                        for detail in child_details
                        if detail.split.error
                    ),
                    None,
                ),
                display_operations=tuple(
                    detail.output_operation for detail in child_details
                ),
            )
            regrouped_rows.append(
                ExcelRowResult(
                    input_row=original_row.input_row,
                    split=shared_split,
                    details=[
                        replace(
                            detail,
                            split=shared_split,
                            child_index=index,
                        )
                        for index, detail in enumerate(
                            child_details,
                            start=1,
                        )
                    ],
                )
            )
        reviewed_rows = regrouped_rows

    workbook = load_workbook(BytesIO(decomposition.source_bytes), data_only=False)
    decomposition_bytes, _ = _write_decomposition(workbook, reviewed_rows)
    return replace(
        decomposition,
        decomposition_bytes=decomposition_bytes,
        rows=reviewed_rows,
        timings=list(decomposition.timings),
    )


def _clone_rows_for_analysis(rows: list[ExcelRowResult]) -> list[ExcelRowResult]:
    cloned_rows: list[ExcelRowResult] = []
    for row in rows:
        details = [
            replace(
                detail,
                result=None,
                error=None,
                repeated_action_trace=None,
            )
            for detail in row.details
        ]
        cloned_rows.append(
            ExcelRowResult(input_row=row.input_row, split=row.split, details=details)
        )
    return cloned_rows


async def _classify_pending_review_actors(
    rows: list[ExcelRowResult],
    deps: Deps,
    sem: asyncio.Semaphore,
) -> None:
    """为审核时新增的歧义动作补做主体判定，但绝不再次执行动作拆解。
    不做去重:每个待审核明细单独判定。"""
    pending_units: list[ExcelDetailResult] = []
    for row in rows:
        for detail in row.details:
            if detail.split.actor == PENDING_REVIEW_ACTOR:
                pending_units.append(detail)

    async def classify_group(detail: ExcelDetailResult) -> None:
        try:
            async with sem:
                actor, source = await classify_operation_actor(
                    detail.operation,
                    deps,
                )
            split = OperationSplit(
                actor=actor,
                operations=(detail.operation,),
                source=f"人工审核确认 + {source}",
            )
        except Exception as exc:
            logger.exception(
                "Reviewed operation actor classification failed: operation=%r",
                detail.operation,
            )
            split = OperationSplit(
                actor="人工",
                operations=(detail.operation,),
                source="人工审核后的主体判定失败",
                needs_review=True,
                error=f"{type(exc).__name__}: {exc}",
            )
        detail.split = split

    await asyncio.gather(
        *(classify_group(detail) for detail in pending_units)
    )
    for row in rows:
        if row.split.actor == PENDING_REVIEW_ACTOR and row.details:
            row.split = row.details[0].split


async def _resolve_parent_weight_groups(
    rows: list[ExcelRowResult],
    deps: Deps,
    sem: asyncio.Semaphore,
) -> dict[int, PartWeightGroupResolution]:
    """按父工序签名解析人工子动作重量组；设备子动作不接收重量上下文。"""
    resolutions = {
        id(row): PartWeightGroupResolution()
        for row in rows
    }
    grouped_rows: dict[tuple, list[ExcelRowResult]] = {}
    for row in rows:
        manual_details = [
            detail
            for detail in row.details
            if detail.effective_actor == "人工"
        ]
        if not manual_details:
            continue
        signature = (
            normalize(row.input_row.operation) or row.input_row.operation,
            tuple(
                (
                    detail.effective_actor,
                    normalize(detail.operation) or detail.operation,
                )
                for detail in row.details
            ),
        )
        grouped_rows.setdefault(signature, []).append(row)

    async def resolve_group(group: list[ExcelRowResult]) -> None:
        representative = group[0]
        representative_manual_details = [
            detail
            for detail in representative.details
            if detail.effective_actor == "人工"
        ]
        async with sem:
            resolution = await resolve_part_weight_groups(
                representative.input_row.operation,
                tuple(
                    detail.operation
                    for detail in representative_manual_details
                ),
                deps,
            )
        contexts = {
            detail.child_index: resolution.contexts[manual_index]
            for manual_index, detail in enumerate(
                representative_manual_details,
                start=1,
            )
            if manual_index in resolution.contexts
        }
        identity_contexts = {
            detail.child_index: resolution.identity_contexts[manual_index]
            for manual_index, detail in enumerate(
                representative_manual_details,
                start=1,
            )
            if manual_index in resolution.identity_contexts
        }
        scoped_resolution = PartWeightGroupResolution(
            contexts=contexts,
            identity_contexts=identity_contexts,
            attempted=resolution.attempted,
        )
        for row in group:
            resolutions[id(row)] = scoped_resolution

    await asyncio.gather(
        *(resolve_group(group) for group in grouped_rows.values())
    )
    return resolutions


@dataclass
class _DetailAnalysisUnit:
    details: list[ExcelDetailResult]
    weight_resolution: PartWeightGroupResolution
    analysis_operation: str
    repeated_group: Optional[RepeatedActionGroup] = None


def _repeated_action_trace(unit: _DetailAnalysisUnit) -> Optional[tuple[str, str, str]]:
    repeated_group = unit.repeated_group
    if repeated_group is None:
        return None
    members = [detail.child_index for detail in unit.details]
    return (
        "RepeatedActionConsistency",
        f"{repeated_group.group_id}: {repeated_group.canonical_operation}",
        f"members={members}",
    )


async def analyze_decomposition_output(
    decomposition: ExcelDecompositionOutput,
    deps: Deps,
    *,
    resolver: Resolver = resolve,
    concurrency: Optional[int] = None,
    on_progress: Optional[ProgressCallback] = None,
) -> ExcelBatchOutput:
    """对已确认的拆解快照执行工时分析，不再次拆解或翻译。"""
    analysis_batch_started = time.perf_counter()
    rows = _clone_rows_for_analysis(decomposition.rows)
    timings = list(decomposition.timings)
    sem = asyncio.Semaphore(max(1, concurrency or settings.CONCURRENCY_LIMIT))
    analysis_started = time.perf_counter()
    await _classify_pending_review_actors(rows, deps, sem)
    parent_weight_resolutions = await _resolve_parent_weight_groups(
        rows,
        deps,
        sem,
    )

    # 相同动作只有在“父重量解析作用域 + 当前重量上下文”也一致时才复用。
    # 这样既保留原有跨行去重，又不会把不同父工序/不同零件的重量串在一起。
    grouped_detail_units: dict[tuple, _DetailAnalysisUnit] = {}
    for row in rows:
        weight_resolution = parent_weight_resolutions[id(row)]
        repeated_resolution = build_repeated_action_groups(
            tuple(
                detail.operation
                if detail.effective_actor == "人工"
                else ""
                for detail in row.details
            )
        )
        repeated_by_child: dict[int, RepeatedActionGroup] = {}
        for candidate in repeated_resolution.groups:
            context_partitions: dict[
                tuple[Optional[int], Optional[int]],
                list[int],
            ] = {}
            for child_index in candidate.child_indexes:
                context = weight_resolution.contexts.get(child_index)
                identity_context = weight_resolution.identity_contexts.get(
                    child_index
                )
                context_scope = (
                    id(context) if context is not None else None,
                    id(identity_context)
                    if identity_context is not None
                    else None,
                )
                context_partitions.setdefault(context_scope, []).append(
                    child_index
                )
            for child_indexes in context_partitions.values():
                if len(child_indexes) > 1:
                    for child_index in child_indexes:
                        repeated_by_child[child_index] = candidate
        for detail in row.details:
            is_manual = detail.effective_actor == "人工"
            numeric_context = (
                weight_resolution.contexts.get(detail.child_index)
                if is_manual
                else None
            )
            part_identity_context = (
                weight_resolution.identity_contexts.get(detail.child_index)
                if is_manual
                else None
            )
            weight_scope = (
                id(weight_resolution)
                if is_manual and weight_resolution.attempted
                else None
            )
            numeric_scope = (
                id(numeric_context)
                if numeric_context is not None
                else None
            )
            identity_scope = (
                id(part_identity_context)
                if part_identity_context is not None
                else None
            )
            repeated_group = repeated_by_child.get(detail.child_index)
            if repeated_group is None:
                analysis_operation = detail.operation
                key = (
                    "exact",
                    detail.effective_actor,
                    normalize(detail.operation) or detail.operation,
                    weight_scope,
                    numeric_scope,
                    identity_scope,
                )
            else:
                analysis_operation = repeated_group.canonical_operation
                key = (
                    "repeated",
                    id(row),
                    detail.effective_actor,
                    repeated_group.group_id,
                    normalize(analysis_operation) or analysis_operation,
                    weight_scope,
                    numeric_scope,
                    identity_scope,
                )
            unit = grouped_detail_units.get(key)
            if unit is None:
                grouped_detail_units[key] = _DetailAnalysisUnit(
                    details=[detail],
                    weight_resolution=weight_resolution,
                    analysis_operation=analysis_operation,
                    repeated_group=repeated_group,
                )
            else:
                unit.details.append(detail)
    detail_units = list(grouped_detail_units.values())

    analysis_completed = 0
    total_details = sum(len(unit.details) for unit in detail_units)

    async def analyze_group(unit: _DetailAnalysisUnit) -> None:
        nonlocal analysis_completed
        group = unit.details
        weight_resolution = unit.weight_resolution
        representative = group[0]
        effective_actor = representative.effective_actor
        is_manual = effective_actor == "人工"
        repeated_trace = _repeated_action_trace(unit)
        if repeated_trace is not None:
            for detail in group:
                detail.repeated_action_trace = repeated_trace
        item_started: Optional[float] = None
        try:
            async with sem:
                item_started = time.perf_counter()
                element = StdsElement(
                    number=representative.input_row.number,
                    operation_des=unit.analysis_operation,
                    line_name=str(representative.input_row.line_name or "").strip(),
                    station_op=str(representative.input_row.station_op or "").strip(),
                    freq=1.0,
                    norm_key=normalize(unit.analysis_operation)
                    or unit.analysis_operation,
                )
                result = await resolve_with_actor(
                    resolver,
                    element,
                    deps,
                    effective_actor,
                    numeric_context=(
                        weight_resolution.contexts.get(
                            representative.child_index
                        )
                        if is_manual
                        else None
                    ),
                    part_identity_context=(
                        weight_resolution.identity_contexts.get(
                            representative.child_index
                        )
                        if is_manual
                        else None
                    ),
                    part_context_resolved=(
                        weight_resolution.attempted
                        if is_manual
                        else False
                    ),
                )
                if effective_actor == "设备" and result.source != Source.MACHINE:
                    # 有效主体是输出硬约束；防止外部 resolver 忽略 hint 后
                    # 把旧人工结果写入设备子工序。
                    result = StdsResult.machine_placeholder(element)
                result_trace = list(result.trace or [])
                if repeated_trace is not None:
                    if not result_trace:
                        result_trace = result_trace_items(result)
                    result_trace = [repeated_trace, *result_trace]
                # 一致性组只共享单次决策；成员数绝不写入 freq 或放大单条工时。
                for detail in group:
                    detail_element = StdsElement(
                        number=detail.input_row.number,
                        operation_des=detail.operation,
                        line_name=str(detail.input_row.line_name or "").strip(),
                        station_op=str(detail.input_row.station_op or "").strip(),
                        freq=1.0,
                        norm_key=normalize(detail.operation) or detail.operation,
                    )
                    detail.result = replace(
                        result,
                        element=detail_element,
                        trace=list(result_trace),
                    )
        except Exception as exc:
            logger.exception(
                "Excel decomposed operation failed: operation=%r rows=%s",
                representative.operation,
                [
                    (detail.input_row.sheet_name, detail.input_row.row_index)
                    for detail in group
                ],
            )
            error = f"{type(exc).__name__}: {exc}"
            for detail in group:
                detail.error = error
        finally:
            item_elapsed_s = (
                time.perf_counter() - item_started if item_started is not None else 0.0
            )
            per_row_elapsed_s = item_elapsed_s / len(group)
            for detail in group:
                analysis_completed += 1
                progress = ExcelProgress(
                    phase=PHASE_ANALYZE,
                    completed_rows=analysis_completed,
                    total_rows=total_details,
                    operation=detail.operation,
                    affected_rows=1,
                    item_elapsed_s=per_row_elapsed_s,
                    total_elapsed_s=(
                        decomposition.total_elapsed_s
                        + time.perf_counter()
                        - analysis_batch_started
                    ),
                    actor=detail.effective_actor,
                    sheet_name=detail.input_row.sheet_name,
                    row_index=detail.input_row.row_index,
                    number=detail.input_row.number,
                    project_name=detail.input_row.project_name,
                    product_model=detail.input_row.product_model,
                    line_name=detail.input_row.line_name,
                    station_op=detail.input_row.station_op,
                    station_description=detail.input_row.station_description,
                    child_index=detail.child_index,
                    child_count=detail.child_count,
                )
                timings.append(progress)
                await _notify_excel_progress(on_progress, progress)

    await asyncio.gather(*(analyze_group(unit) for unit in detail_units))
    analysis_elapsed_s = time.perf_counter() - analysis_started
    final_workbook = load_workbook(
        BytesIO(decomposition.source_bytes),
        data_only=False,
    )
    output_bytes, detail_sheet_name = _write_results(final_workbook, rows)
    total_elapsed_s = (
        decomposition.total_elapsed_s
        + time.perf_counter()
        - analysis_batch_started
    )
    return ExcelBatchOutput(
        output_bytes=output_bytes,
        output_filename=build_output_filename(decomposition.source_name),
        decomposition_bytes=decomposition.decomposition_bytes,
        decomposition_filename=decomposition.decomposition_filename,
        rows=rows,
        timings=timings,
        total_elapsed_s=total_elapsed_s,
        decompose_elapsed_s=decomposition.decompose_elapsed_s,
        analysis_elapsed_s=analysis_elapsed_s,
        detail_sheet_name=detail_sheet_name,
    )


async def analyze_excel_bytes(
    excel_bytes: bytes,
    source_name: str,
    deps: Deps,
    *,
    resolver: Resolver = resolve,
    decomposer: Decomposer = decompose_operation,
    translator: OutputTranslator = translate_operation_for_output,
    concurrency: Optional[int] = None,
    on_progress: Optional[ProgressCallback] = None,
) -> ExcelBatchOutput:
    """默认自动模式：拆解和翻译完成后直接进入工时分析。"""
    decomposition = await decompose_excel_bytes(
        excel_bytes,
        source_name,
        deps,
        decomposer=decomposer,
        translator=translator,
        concurrency=concurrency,
        on_progress=on_progress,
    )
    return await analyze_decomposition_output(
        decomposition,
        deps,
        resolver=resolver,
        concurrency=concurrency,
        on_progress=on_progress,
    )
