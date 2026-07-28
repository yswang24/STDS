"""零件名称到单重的精确/语义检索。

重量表仅在启动时读取；语义向量在首次需要模糊检索时按需构建并保存在内存。
完全重复的定义严格为：零件号、英文名称、中文名称、重量四项全部一致。
"""
from __future__ import annotations

import asyncio
import logging
import math
import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Optional, Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from stds.config.settings import settings
from stds.retrieval.embed import EmbedBackend, get_embed_backend

logger = logging.getLogger("stds.part_weight")

_REQUIRED_HEADERS = {
    "part_no": "Part No.",
    "english_name": "English Description",
    "chinese_name": "Chinese Description",
    "weight": "零件单重(KG)",
}
_MATCH_NORMALIZE_RE = re.compile(r"[\W_]+", re.UNICODE)


@dataclass(frozen=True)
class PartWeightSource:
    sheet: str
    row: int
    cell: str


@dataclass
class PartWeightRecord:
    part_no: str
    english_name: str
    chinese_name: str
    weight_kg: Optional[float]
    sources: list[PartWeightSource] = field(default_factory=list)
    _weight_key: tuple[str, object] = field(
        default=("none", ""),
        repr=False,
        compare=False,
    )

    def __post_init__(self) -> None:
        if self._weight_key == ("none", "") and self.weight_kg is not None:
            self._weight_key = (
                "number",
                Decimal(str(self.weight_kg)).normalize(),
            )

@dataclass(frozen=True)
class PartWeightMatch:
    query: str
    matched_name: str
    part_no: str
    weight_kg: float
    similarity: float
    match_type: str
    sources: tuple[PartWeightSource, ...]

    @property
    def source_label(self) -> str:
        if not self.sources:
            return ""
        source = self.sources[0]
        return f"{source.sheet}!{source.cell}"


def normalize_part_name(value: object) -> str:
    """匹配用名称归一化；不参与严格去重键。"""
    text = unicodedata.normalize("NFKC", str(value or "")).casefold().strip()
    return _MATCH_NORMALIZE_RE.sub("", text)


def _part_no_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _weight_value_and_key(value: object) -> tuple[Optional[float], tuple[str, object]]:
    """返回可计算公斤值和严格去重使用的稳定键。"""
    if value is None or isinstance(value, bool):
        return None, ("none", "")
    if isinstance(value, (int, float, Decimal)):
        try:
            decimal_value = Decimal(str(value))
        except InvalidOperation:
            return None, ("text", str(value).strip())
        if not decimal_value.is_finite():
            return None, ("text", str(value).strip())
        return float(decimal_value), ("number", decimal_value.normalize())
    text = str(value).strip()
    if not text:
        return None, ("text", "")
    try:
        decimal_value = Decimal(text)
    except InvalidOperation:
        return None, ("text", text)
    if not decimal_value.is_finite():
        return None, ("text", text)
    return float(decimal_value), ("number", decimal_value.normalize())


def load_part_weight_records(
    path: Union[str, Path],
) -> list[PartWeightRecord]:
    """从工作簿加载并按四字段严格去重，保留全部来源位置。"""
    source_path = Path(path)
    if not source_path.is_file():
        logger.warning("零件重量表不存在，重量增强已关闭: %s", source_path)
        return []

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    deduped: dict[tuple, PartWeightRecord] = {}
    try:
        for worksheet in workbook.worksheets:
            header_values = [
                str(cell.value).strip() if cell.value is not None else ""
                for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
            ]
            header_index = {
                name: header_values.index(label)
                for name, label in _REQUIRED_HEADERS.items()
                if label in header_values
            }
            if len(header_index) != len(_REQUIRED_HEADERS):
                logger.info("跳过不符合重量表结构的工作表: %s", worksheet.title)
                continue

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True),
                start=2,
            ):
                part_no = _part_no_text(row[header_index["part_no"]])
                english_name = str(
                    row[header_index["english_name"]] or ""
                ).strip()
                chinese_name = str(
                    row[header_index["chinese_name"]] or ""
                ).strip()
                raw_weight = row[header_index["weight"]]
                if not part_no and not english_name and not chinese_name and raw_weight is None:
                    continue
                weight_kg, weight_key = _weight_value_and_key(raw_weight)
                key = (
                    part_no,
                    english_name,
                    chinese_name,
                    weight_key,
                )
                source = PartWeightSource(
                    sheet=worksheet.title,
                    row=row_number,
                    cell=(
                        f"{get_column_letter(header_index['weight'] + 1)}"
                        f"{row_number}"
                    ),
                )
                existing = deduped.get(key)
                if existing is not None:
                    existing.sources.append(source)
                    continue
                deduped[key] = PartWeightRecord(
                    part_no=part_no,
                    english_name=english_name,
                    chinese_name=chinese_name,
                    weight_kg=weight_kg,
                    sources=[source],
                    _weight_key=weight_key,
                )
    finally:
        workbook.close()
    return list(deduped.values())


def _cosine(a: list[float], b: list[float]) -> float:
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = math.sqrt(sum(x * x for x in a)) or 1e-9
    norm_b = math.sqrt(sum(x * x for x in b)) or 1e-9
    return dot / (norm_a * norm_b)


class PartWeightIndex:
    def __init__(
        self,
        records: list[PartWeightRecord],
        embed_backend: Optional[EmbedBackend] = None,
        *,
        embed_backend_factory: Callable[[], EmbedBackend] = get_embed_backend,
        similarity_threshold: float = settings.PART_WEIGHT_SIMILARITY_THRESHOLD,
        similarity_margin: float = settings.PART_WEIGHT_SIMILARITY_MARGIN,
    ):
        self.records = list(records)
        self._embed = embed_backend
        self._embed_backend_factory = embed_backend_factory
        self.similarity_threshold = similarity_threshold
        self.similarity_margin = similarity_margin
        self._exact: dict[str, list[PartWeightRecord]] = {}
        for record in self.records:
            keys = {
                normalize_part_name(name)
                for name in (record.chinese_name, record.english_name)
                if normalize_part_name(name)
            }
            for key in keys:
                self._exact.setdefault(key, []).append(record)
        self._semantic_entries = [
            (record, name)
            for record in self.records
            for name in (record.chinese_name, record.english_name)
            if normalize_part_name(name)
        ]
        self._semantic_vectors: Optional[list[list[float]]] = None
        self._build_lock: Optional[asyncio.Lock] = None

    @classmethod
    def from_xlsx(
        cls,
        path: Union[str, Path],
        embed_backend: Optional[EmbedBackend] = None,
        **kwargs,
    ) -> "PartWeightIndex":
        return cls(
            load_part_weight_records(path),
            embed_backend=embed_backend,
            **kwargs,
        )

    @property
    def available(self) -> bool:
        return bool(self.records)

    def _consistent_match(
        self,
        query: str,
        records: list[PartWeightRecord],
        *,
        similarity: float,
        match_type: str,
        matched_name: Optional[str] = None,
    ) -> Optional[PartWeightMatch]:
        valid = [
            record
            for record in records
            if record.weight_kg is not None and record.weight_kg > 0
        ]
        weights = {record._weight_key for record in valid}
        if not valid or len(weights) != 1:
            return None
        record = valid[0]
        resolved_name = matched_name or record.chinese_name or record.english_name
        sources = tuple(
            source
            for candidate in valid
            for source in candidate.sources
        )
        return PartWeightMatch(
            query=query,
            matched_name=resolved_name,
            part_no=record.part_no,
            weight_kg=float(record.weight_kg),
            similarity=float(similarity),
            match_type=match_type,
            sources=sources,
        )

    def exact_match(self, query: str) -> Optional[PartWeightMatch]:
        key = normalize_part_name(query)
        if not key:
            return None
        records = self._exact.get(key, [])
        matched_name = next(
            (
                name
                for record in records
                for name in (record.chinese_name, record.english_name)
                if normalize_part_name(name) == key
            ),
            query,
        )
        return self._consistent_match(
            query,
            records,
            similarity=1.0,
            match_type="exact",
            matched_name=matched_name,
        )

    async def _ensure_semantic_vectors(self) -> None:
        if self._semantic_vectors is not None:
            return
        if self._build_lock is None:
            self._build_lock = asyncio.Lock()
        async with self._build_lock:
            if self._semantic_vectors is not None:
                return
            if self._embed is None:
                self._embed = self._embed_backend_factory()
            texts = [name for _, name in self._semantic_entries]
            self._semantic_vectors = (
                await asyncio.to_thread(self._embed.embed, texts)
                if texts
                else []
            )

    async def semantic_match(
        self,
        query: str,
        *,
        top_k: int = 3,
    ) -> Optional[PartWeightMatch]:
        normalized = normalize_part_name(query)
        if not normalized or not self._semantic_entries:
            return None
        # 很短的纯英文缩写只能精确匹配，避免 BDU/CMU 一类缩写误召回。
        if re.fullmatch(r"[a-z0-9]{1,4}", normalized):
            return None

        await self._ensure_semantic_vectors()
        if self._embed is None or not self._semantic_vectors:
            return None
        query_vector = await asyncio.to_thread(self._embed.embed_one, query)
        scored = sorted(
            (
                (_cosine(query_vector, vector), record, name)
                for (record, name), vector in zip(
                    self._semantic_entries,
                    self._semantic_vectors,
                )
            ),
            key=lambda item: item[0],
            reverse=True,
        )
        if not scored:
            return None

        top_score, _, top_name = scored[0]
        top_name_key = normalize_part_name(top_name)
        top_group = []
        seen_record_ids = set()
        for _, record, name in scored:
            if (
                normalize_part_name(name) == top_name_key
                and id(record) not in seen_record_ids
            ):
                top_group.append(record)
                seen_record_ids.add(id(record))
        top_group_ids = {id(record) for record in top_group}
        distinct = [
            (score, record, name)
            for score, record, name in scored
            if id(record) not in top_group_ids
        ][: max(1, top_k - 1)]
        second_score = distinct[0][0] if distinct else -1.0
        if top_score < self.similarity_threshold:
            return None
        if top_score - second_score < self.similarity_margin:
            return None
        return self._consistent_match(
            query,
            top_group,
            similarity=top_score,
            match_type="semantic",
            matched_name=top_name,
        )

    async def match(self, query: str) -> Optional[PartWeightMatch]:
        exact = self.exact_match(query)
        if exact is not None:
            return exact
        return await self.semantic_match(query)


def load_part_weight_index(
    path: Optional[str] = None,
    *,
    embed_backend: Optional[EmbedBackend] = None,
) -> PartWeightIndex:
    return PartWeightIndex.from_xlsx(
        path or settings.PART_WEIGHT_XLSX_PATH,
        embed_backend=embed_backend,
    )
