"""从随主输入上传的 Excel 中加载 Chartcode 与参数选择经验。"""
from __future__ import annotations

import hashlib
import io
import math
import re
import unicodedata
from dataclasses import dataclass, replace
from pathlib import Path
from typing import BinaryIO, Optional, Union

from openpyxl import load_workbook

from stds.engine.decision_codec import (
    canonicalize_decision,
    decode_strict_with_trace,
)
from stds.engine.formula import evaluate
from stds.experience.common_chart import normalize_common_keyword
from stds.experience.common_index import CommonChartSemanticIndex
from stds.experience.index import (
    ExperienceIndex,
    normalize_chartcode,
    normalize_operation,
)
from stds.experience.models import (
    CommonChartEntry,
    CommonChartKind,
    ExperienceEntry,
    ExperienceIssue,
    ExperienceLoadResult,
    ParameterExperienceEntry,
)
from stds.retrieval.embed import EmbedBackend

WorkbookSource = Union[str, Path, bytes, bytearray, memoryview, BinaryIO]

_CHART_SHEET = "chartcode选择经验"
_PARAMETER_SHEET = "参数选择经验"
_COMMON_SHEET = "Common_Chart"
_EST_CHARTCODES = {"ESTC00", "ESTV00"}
_VARIABLE_MARKER = re.compile(
    r"(?:参数\s*)?[VＶvｖ]\s*(\d+)\s*[：:]",
    re.UNICODE,
)
_MEASUREMENT_RE = re.compile(
    r"(?<![\d.])(\d+(?:\.\d+)?)\s*"
    r"(cm|厘米|m|米|in|英寸|ft|度|°)(?![a-z])",
    re.IGNORECASE,
)
_PLAIN_NUMBER_RE = re.compile(r"(?<![\d.])\d+(?:\.\d+)?(?![\d.])")


@dataclass(frozen=True)
class _ParameterCandidate:
    entry: ParameterExperienceEntry
    explicit_id: bool
    rule_key: str


def _source_bytes(source: WorkbookSource) -> tuple[bytes, str]:
    if isinstance(source, (str, Path)):
        path = Path(source)
        return path.read_bytes(), path.name
    if isinstance(source, (bytes, bytearray, memoryview)):
        return bytes(source), ""
    if not hasattr(source, "read"):
        raise TypeError("source 必须是 path、bytes、BytesIO 或二进制文件对象")

    inferred_name = Path(str(getattr(source, "name", ""))).name
    original_position: Optional[int] = None
    try:
        original_position = source.tell()
    except (AttributeError, OSError):
        pass
    try:
        try:
            source.seek(0)
        except (AttributeError, OSError):
            pass
        data = source.read()
    finally:
        if original_position is not None:
            try:
                source.seek(original_position)
            except (AttributeError, OSError):
                pass
    if not isinstance(data, (bytes, bytearray, memoryview)):
        raise TypeError("文件对象必须以二进制方式读取")
    return bytes(data), inferred_name


def _header_map(worksheet) -> dict[str, int]:
    try:
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return {}
    return {
        str(value).strip(): index
        for index, value in enumerate(header)
        if value is not None and str(value).strip()
    }


def _header_index(headers: dict[str, int], *aliases: str) -> Optional[int]:
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def _text(value: object) -> str:
    return str(value or "").strip()


def _cell(row: tuple, index: Optional[int]) -> object:
    return row[index] if index is not None and index < len(row) else None


def split_variable_hints(parameter_text: str) -> dict[int, str]:
    """将“参数Vn：...”拆成当前变量可直接检索的提示。"""
    matches = list(_VARIABLE_MARKER.finditer(parameter_text or ""))
    hints: dict[int, str] = {}
    for index, marker in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(parameter_text)
        hint = parameter_text[marker.end():end].strip().rstrip("；;").strip()
        variable_number = int(marker.group(1))
        if not hint:
            continue
        if variable_number in hints:
            hints[variable_number] = f"{hints[variable_number]}\n{hint}"
        else:
            hints[variable_number] = hint
    return hints


def _derived_experience_id(operation_key: str, chartcode: str) -> str:
    return f"auto:{operation_key}|{normalize_chartcode(chartcode)}"


def _canonical_parameter_rule(value: str) -> str:
    text = unicodedata.normalize("NFKC", value or "").strip()
    return re.sub(r"\s+", " ", text)


def _canonical_measurement(number: str, unit: str) -> tuple[str, float]:
    value = float(number)
    normalized_unit = unit.casefold()
    if normalized_unit in {"cm", "厘米"}:
        return "length_m", value / 100.0
    if normalized_unit in {"m", "米"}:
        return "length_m", value
    if normalized_unit in {"in", "英寸"}:
        return "length_m", value * 0.0254
    if normalized_unit == "ft":
        return "length_m", value * 0.3048
    return "angle_deg", value


def _chart_has_variable(chart, variable_number: int) -> Optional[bool]:
    """目标图表可检查时，确认 Vn 至少存在一组真实候选。"""
    options_by_node = getattr(chart, "options", None)
    if not hasattr(options_by_node, "items"):
        return None
    return any(
        candidate_variable == variable_number and bool(options)
        for (candidate_variable, _), options in options_by_node.items()
    )


def _invalid_measurements(
    hint: str,
    chart,
    variable_number: int,
) -> list[str]:
    """发现经验默认值与当前图表候选明显不一致的单位/数值。"""
    requested = [
        (match.group(0), *_canonical_measurement(match.group(1), match.group(2)))
        for match in _MEASUREMENT_RE.finditer(hint)
    ]
    if not requested:
        return []

    options_by_node = getattr(chart, "options", None)
    if not hasattr(options_by_node, "items"):
        return []
    candidates = [
        option
        for (candidate_variable, _), options in options_by_node.items()
        if candidate_variable == variable_number
        for option in options
    ]
    candidate_measurements = [
        _canonical_measurement(match.group(1), match.group(2))
        for candidate in candidates
        for text in (
            str(candidate.description or ""),
            str(candidate.metric_abbrev or ""),
        )
        for match in _MEASUREMENT_RE.finditer(text)
    ]
    candidate_plain_numbers = {
        float(number)
        for candidate in candidates
        for number in _PLAIN_NUMBER_RE.findall(
            str(candidate.description or "")
        )
    }

    invalid = []
    for raw, kind, value in requested:
        if value == 0 and 0.0 in candidate_plain_numbers:
            continue
        tolerance = 0.005 if kind == "length_m" else 0.01
        if not any(
            candidate_kind == kind
            and abs(candidate_value - value) <= tolerance
            for candidate_kind, candidate_value in candidate_measurements
        ):
            # 角度图表的候选常只写“45/90/180”，没有再写“度”。
            if kind == "angle_deg" and value in candidate_plain_numbers:
                continue
            invalid.append(raw)
    return invalid


def _finite_float(value: object) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def _load_common_entries(
    workbook,
    charts: dict,
    chart_lookup: dict[str, str],
    issues: list[ExperienceIssue],
) -> list[CommonChartEntry]:
    """解析并逐行校验可选的 Common_Chart 工作表。"""
    if _COMMON_SHEET not in workbook.sheetnames:
        issues.append(ExperienceIssue(
            "warning",
            "missing_sheet",
            f"缺少工作表“{_COMMON_SHEET}”，Common Chart 快速路径不可用",
            _COMMON_SHEET,
        ))
        return []

    worksheet = workbook[_COMMON_SHEET]
    headers = _header_map(worksheet)
    operation_col = _header_index(headers, "操作内容")
    decision_col = _header_index(headers, "决策描述")
    chartcode_col = _header_index(headers, "动作代码", "Chartcode")
    cv_col = _header_index(
        headers,
        "增值/非增值(C/V)",
        "增值非增值",
        "C/V",
    )
    frequency_col = _header_index(headers, "频率")
    time_col = _header_index(headers, "时间")
    required = {
        "操作内容": operation_col,
        "决策描述": decision_col,
        "动作代码": chartcode_col,
        "增值/非增值(C/V)": cv_col,
        "频率": frequency_col,
        "时间": time_col,
    }
    missing = [name for name, index in required.items() if index is None]
    if missing:
        issues.append(ExperienceIssue(
            "error",
            "missing_header",
            f"Common_Chart 缺少列：{'、'.join(missing)}",
            _COMMON_SHEET,
        ))
        return []

    keyword_columns: list[tuple[str, int]] = []
    for number in range(1, 9):
        index = _header_index(
            headers,
            f"关键词描述{number}",
            f"关键词{number}",
        )
        if index is not None:
            keyword_columns.append((f"关键词描述{number}", index))
    if not keyword_columns:
        issues.append(ExperienceIssue(
            "error",
            "missing_header",
            "Common_Chart 缺少“关键词描述1..8”列",
            _COMMON_SHEET,
        ))
        return []

    entries: list[CommonChartEntry] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        operation_label = _text(_cell(row, operation_col))
        raw_decision = _text(_cell(row, decision_col))
        raw_chartcode = _text(_cell(row, chartcode_col))
        raw_cv = _text(_cell(row, cv_col)).upper()
        raw_frequency = _cell(row, frequency_col)
        raw_time = _cell(row, time_col)
        raw_keywords = [
            (_text(_cell(row, index)), field_name)
            for field_name, index in keyword_columns
        ]
        if not any((
            operation_label,
            raw_decision,
            raw_chartcode,
            raw_cv,
            _text(raw_frequency),
            _text(raw_time),
            *(keyword for keyword, _ in raw_keywords),
        )):
            continue

        if not operation_label:
            issues.append(ExperienceIssue(
                "warning",
                "empty_common_operation",
                "操作内容为空，Common_Chart 行已忽略",
                _COMMON_SHEET,
                row_number,
                "操作内容",
            ))
            continue

        chartcode_key = normalize_chartcode(raw_chartcode)
        chartcode = chart_lookup.get(chartcode_key)
        if chartcode is None:
            issues.append(ExperienceIssue(
                "warning",
                "invalid_common_chartcode",
                f"Chartcode“{raw_chartcode}”不在当前图表库中，Common_Chart 行已忽略",
                _COMMON_SHEET,
                row_number,
                "动作代码",
            ))
            continue
        chart = charts[chartcode]

        if raw_cv not in {"C", "V"}:
            issues.append(ExperienceIssue(
                "warning",
                "invalid_common_cv",
                "增值/非增值必须是 C 或 V，Common_Chart 行已忽略",
                _COMMON_SHEET,
                row_number,
                "增值/非增值(C/V)",
            ))
            continue
        if chartcode_key == "ESTC00":
            expected_cv = "C"
        elif chartcode_key == "ESTV00":
            expected_cv = "V"
        elif hasattr(chart, "value_added"):
            expected_cv = "C" if bool(chart.value_added) else "V"
        else:
            expected_cv = raw_cv
        if raw_cv != expected_cv:
            issues.append(ExperienceIssue(
                "warning",
                "common_cv_mismatch",
                (
                    f"C/V={raw_cv} 与 Chartcode“{chartcode}”的"
                    f"{expected_cv}属性不一致，Common_Chart 行已忽略"
                ),
                _COMMON_SHEET,
                row_number,
                "增值/非增值(C/V)",
            ))
            continue

        frequency = _finite_float(raw_frequency)
        if frequency is None or not math.isclose(
            frequency,
            1.0,
            rel_tol=0.0,
            abs_tol=1e-9,
        ):
            issues.append(ExperienceIssue(
                "warning",
                "invalid_common_frequency",
                "Common_Chart 频率必须为 1（拆分结果按单次计时），该行已忽略",
                _COMMON_SHEET,
                row_number,
                "频率",
            ))
            continue

        source_time_s = _finite_float(raw_time)
        if source_time_s is None or source_time_s < 0:
            issues.append(ExperienceIssue(
                "warning",
                "invalid_common_time",
                "时间必须是有限的非负数，Common_Chart 行已忽略",
                _COMMON_SHEET,
                row_number,
                "时间",
            ))
            continue

        keywords: list[str] = []
        normalized_keywords: list[str] = []
        seen_keywords: set[str] = set()
        for keyword, field_name in raw_keywords:
            if not keyword:
                continue
            normalized_keyword = normalize_common_keyword(keyword)
            if not normalized_keyword:
                issues.append(ExperienceIssue(
                    "warning",
                    "invalid_common_keyword",
                    (
                        f"关键词“{keyword}”过短：纯中文至少 2 字，"
                        "其他形式至少 3 字符，已忽略该关键词"
                    ),
                    _COMMON_SHEET,
                    row_number,
                    field_name,
                ))
                continue
            if normalized_keyword in seen_keywords:
                continue
            seen_keywords.add(normalized_keyword)
            keywords.append(keyword)
            normalized_keywords.append(normalized_keyword)
        if not keywords:
            issues.append(ExperienceIssue(
                "warning",
                "no_valid_common_keyword",
                "没有可用关键词；该行仍可通过操作内容参与语义检索",
                _COMMON_SHEET,
                row_number,
                "关键词描述1",
            ))

        if chartcode_key in _EST_CHARTCODES:
            if source_time_s <= 0:
                issues.append(ExperienceIssue(
                    "warning",
                    "invalid_fixed_time",
                    "EST 固定时间必须大于 0，Common_Chart 行已忽略",
                    _COMMON_SHEET,
                    row_number,
                    "时间",
                ))
                continue
            time_s = source_time_s / frequency
            decision = f"{time_s:g}S"
            kind = CommonChartKind.FIXED_TIME
            values: dict[int, float] = {}
        else:
            if not raw_decision:
                issues.append(ExperienceIssue(
                    "warning",
                    "empty_common_decision",
                    "普通 Chartcode 的决策描述为空，Common_Chart 行已忽略",
                    _COMMON_SHEET,
                    row_number,
                    "决策描述",
                ))
                continue
            decision = canonicalize_decision(raw_decision)
            try:
                values, _ = decode_strict_with_trace(chart, decision)
                time_s = float(evaluate(chart, values))
            except Exception as exc:
                issues.append(ExperienceIssue(
                    "warning",
                    "invalid_common_decision",
                    f"决策描述无法严格解码：{exc}，Common_Chart 行已忽略",
                    _COMMON_SHEET,
                    row_number,
                    "决策描述",
                ))
                continue
            if not math.isfinite(time_s) or time_s < 0:
                issues.append(ExperienceIssue(
                    "warning",
                    "invalid_common_formula_time",
                    "公式重算结果不是有限的非负数，Common_Chart 行已忽略",
                    _COMMON_SHEET,
                    row_number,
                    "决策描述",
                ))
                continue
            if not math.isclose(
                time_s,
                source_time_s / frequency,
                rel_tol=0.0,
                abs_tol=0.005,
            ):
                issues.append(ExperienceIssue(
                    "warning",
                    "common_time_mismatch",
                    (
                        f"上传时间 {source_time_s:g}s 与公式重算单次时间 "
                        f"{time_s:g}s 不一致；运行时以公式重算值为准"
                    ),
                    _COMMON_SHEET,
                    row_number,
                    "时间",
                ))
            kind = CommonChartKind.FORMULA

        entries.append(CommonChartEntry(
            operation_label=operation_label,
            normalized_operation=normalize_operation(operation_label),
            chartcode=chartcode,
            decision=decision,
            cv=raw_cv,
            frequency=frequency,
            source_time_s=source_time_s,
            time_s=time_s,
            keywords=tuple(keywords),
            normalized_keywords=tuple(normalized_keywords),
            row=row_number,
            kind=kind,
            values=dict(values),
        ))

    # 保留歧义项供运行时基于更长关键词消歧，但在上传阶段显式告警。
    by_keyword: dict[str, list[CommonChartEntry]] = {}
    for entry in entries:
        for keyword in entry.normalized_keywords:
            by_keyword.setdefault(keyword, []).append(entry)
    for keyword, owners in sorted(by_keyword.items()):
        signatures = {entry.output_signature for entry in owners}
        if len(signatures) <= 1:
            continue
        issues.append(ExperienceIssue(
            "warning",
            "ambiguous_common_keyword",
            (
                f"关键词“{keyword}”对应多个不同结果；运行时同优先级命中"
                "时将回退到后续评估路径"
            ),
            _COMMON_SHEET,
        ))
    return entries


def _empty_result(
    digest: str,
    source_name: str,
    issues: list[ExperienceIssue],
    embed_backend: Optional[EmbedBackend],
) -> ExperienceLoadResult:
    common_index = CommonChartSemanticIndex(
        (),
        embed_backend=embed_backend,
    )
    return ExperienceLoadResult(
        index=ExperienceIndex(
            [],
            digest=digest,
            source_name=source_name,
            embed_backend=embed_backend,
        ),
        issues=tuple(issues),
        digest=digest,
        common_index=common_index,
    )


def load_experience_workbook(
    source: WorkbookSource,
    charts: dict,
    *,
    source_name: str = "",
    embed_backend: Optional[EmbedBackend] = None,
) -> ExperienceLoadResult:
    """加载经验工作簿；无效行只产生 warning，不影响其他有效行。"""
    raw, inferred_name = _source_bytes(source)
    digest = hashlib.sha256(raw).hexdigest()
    resolved_source_name = source_name or inferred_name
    issues: list[ExperienceIssue] = []
    try:
        workbook = load_workbook(
            io.BytesIO(raw),
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        issues.append(ExperienceIssue(
            severity="error",
            code="invalid_workbook",
            message=f"无法读取经验工作簿: {exc}",
        ))
        return _empty_result(digest, resolved_source_name, issues, embed_backend)

    chart_lookup = {
        normalize_chartcode(chartcode): chartcode
        for chartcode in charts
        if normalize_chartcode(chartcode)
    }
    entries: list[ExperienceEntry] = []
    parameter_entries: list[ParameterExperienceEntry] = []
    common_entries: list[CommonChartEntry] = []
    try:
        if _CHART_SHEET not in workbook.sheetnames:
            issues.append(ExperienceIssue(
                severity="error",
                code="missing_sheet",
                message=f"缺少工作表“{_CHART_SHEET}”",
                sheet=_CHART_SHEET,
            ))
            return _empty_result(
                digest, resolved_source_name, issues, embed_backend,
            )

        chart_sheet = workbook[_CHART_SHEET]
        chart_headers = _header_map(chart_sheet)
        operation_col = chart_headers.get("操作内容")
        code_col = chart_headers.get("动作代码")
        if code_col is None:
            # 兼容 V1.2：该列虽名为“参数选择”，实际内容是 Chartcode。
            code_col = chart_headers.get("参数选择")
        id_col = chart_headers.get("经验ID")
        if operation_col is None or code_col is None:
            issues.append(ExperienceIssue(
                severity="error",
                code="missing_header",
                message="chartcode选择经验缺少“操作内容”或“动作代码/参数选择”列",
                sheet=_CHART_SHEET,
            ))
            return _empty_result(
                digest, resolved_source_name, issues, embed_backend,
            )

        provisional: list[ExperienceEntry] = []
        for row_number, row in enumerate(
            chart_sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            operation_label = _text(_cell(row, operation_col))
            operation_key = normalize_operation(operation_label)
            raw_chartcode = _text(_cell(row, code_col))
            chartcode = chart_lookup.get(normalize_chartcode(raw_chartcode))
            if not operation_key and not raw_chartcode:
                continue
            if not operation_key:
                issues.append(ExperienceIssue(
                    "warning", "empty_operation", "操作内容为空，已忽略",
                    _CHART_SHEET, row_number, "操作内容",
                ))
                continue
            if chartcode is None:
                issues.append(ExperienceIssue(
                    "warning", "invalid_chartcode",
                    f"Chartcode“{raw_chartcode}”不在当前图表库中，已忽略",
                    _CHART_SHEET, row_number, "动作代码",
                ))
                continue
            explicit_id = _text(_cell(row, id_col))
            experience_id = explicit_id or _derived_experience_id(
                operation_key,
                chartcode,
            )
            provisional.append(ExperienceEntry(
                experience_id=experience_id,
                operation_label=operation_label,
                normalized_operation=operation_key,
                chartcode=chartcode,
                chart_row=row_number,
            ))

        # 同一个显式经验 ID 不可指向不同的动作或图表，否则全部禁用。
        signatures_by_id: dict[str, set[tuple[str, str]]] = {}
        for entry in provisional:
            signatures_by_id.setdefault(entry.experience_id, set()).add(
                (entry.normalized_operation, normalize_chartcode(entry.chartcode))
            )
        conflicted_ids = {
            experience_id
            for experience_id, signatures in signatures_by_id.items()
            if len(signatures) > 1
        }
        for experience_id in sorted(conflicted_ids):
            issues.append(ExperienceIssue(
                "warning", "conflicting_experience_id",
                f"经验ID“{experience_id}”对应多个动作或Chartcode，相关行已禁用",
                _CHART_SHEET,
            ))

        # 完全相同的经验行保留第一条，防止重复内容制造虚假歧义。
        seen: set[tuple[str, str, str]] = set()
        for entry in provisional:
            identity = (
                entry.experience_id,
                entry.normalized_operation,
                normalize_chartcode(entry.chartcode),
            )
            if entry.experience_id in conflicted_ids or identity in seen:
                continue
            seen.add(identity)
            entries.append(entry)

        if _PARAMETER_SHEET in workbook.sheetnames:
            parameter_sheet = workbook[_PARAMETER_SHEET]
            parameter_headers = _header_map(parameter_sheet)
            parameter_operation_col = parameter_headers.get("操作内容")
            parameter_code_col = parameter_headers.get("动作代码")
            parameter_text_col = parameter_headers.get("参数选择经验")
            parameter_id_col = parameter_headers.get("经验ID")
            if (
                parameter_operation_col is None
                or parameter_code_col is None
                or parameter_text_col is None
            ):
                issues.append(ExperienceIssue(
                    "error", "missing_header",
                    "参数选择经验缺少“操作内容”“动作代码”或“参数选择经验”列",
                    _PARAMETER_SHEET,
                ))
            else:
                provisional_parameters: list[_ParameterCandidate] = []
                for row_number, row in enumerate(
                    parameter_sheet.iter_rows(min_row=2, values_only=True),
                    start=2,
                ):
                    operation_label = _text(
                        _cell(row, parameter_operation_col)
                    )
                    operation_key = normalize_operation(operation_label)
                    raw_chartcode = _text(_cell(row, parameter_code_col))
                    chartcode = chart_lookup.get(
                        normalize_chartcode(raw_chartcode)
                    )
                    parameter_text = _text(_cell(row, parameter_text_col))
                    explicit_id = _text(_cell(row, parameter_id_col))
                    if not operation_key and not raw_chartcode and not parameter_text:
                        continue
                    if not operation_key:
                        issues.append(ExperienceIssue(
                            "warning", "empty_operation",
                            "操作内容为空，参数经验已忽略",
                            _PARAMETER_SHEET, row_number, "操作内容",
                        ))
                        continue
                    if chartcode is None:
                        issues.append(ExperienceIssue(
                            "warning", "invalid_chartcode",
                            f"Chartcode“{raw_chartcode}”不在当前图表库中，已忽略",
                            _PARAMETER_SHEET, row_number, "动作代码",
                        ))
                        continue
                    if not parameter_text:
                        issues.append(ExperienceIssue(
                            "warning", "empty_parameter_experience",
                            "参数选择经验为空，已忽略",
                            _PARAMETER_SHEET,
                            row_number,
                            "参数选择经验",
                        ))
                        continue
                    experience_id = explicit_id or _derived_experience_id(
                        operation_key,
                        chartcode,
                    )
                    variable_hints = split_variable_hints(parameter_text)
                    if not variable_hints:
                        issues.append(ExperienceIssue(
                            "warning",
                            "unparsed_parameter_experience",
                            "参数选择经验未解析出任何“参数Vn：...”提示，已忽略",
                            _PARAMETER_SHEET,
                            row_number,
                            "参数选择经验",
                        ))
                        continue

                    chart = charts[chartcode]
                    valid_hints = dict(variable_hints)
                    for variable_number, hint in variable_hints.items():
                        variable_exists = _chart_has_variable(
                            chart,
                            variable_number,
                        )
                        if variable_exists is False:
                            valid_hints.pop(variable_number, None)
                            issues.append(ExperienceIssue(
                                "warning",
                                "invalid_parameter_variable",
                                (
                                    f"参数V{variable_number}不在当前图表"
                                    "候选中，该变量经验已禁用"
                                ),
                                _PARAMETER_SHEET,
                                row_number,
                                f"V{variable_number}",
                            ))
                            continue
                        invalid_values = _invalid_measurements(
                            hint,
                            chart,
                            variable_number,
                        )
                        if not invalid_values:
                            continue
                        valid_hints.pop(variable_number, None)
                        issues.append(ExperienceIssue(
                            "warning",
                            "invalid_parameter_measurement",
                            (
                                f"参数V{variable_number}中的"
                                f"{'、'.join(invalid_values)}"
                                "不在当前图表候选中，该变量经验已禁用"
                            ),
                            _PARAMETER_SHEET,
                            row_number,
                            f"V{variable_number}",
                        ))
                    if not valid_hints:
                        issues.append(ExperienceIssue(
                            "warning",
                            "no_valid_parameter_experience",
                            "参数选择经验校验后没有可用 Vn，整条记录已忽略",
                            _PARAMETER_SHEET,
                            row_number,
                            "参数选择经验",
                        ))
                        continue

                    provisional_parameters.append(_ParameterCandidate(
                        entry=ParameterExperienceEntry(
                            experience_id=experience_id,
                            operation_label=operation_label,
                            normalized_operation=operation_key,
                            chartcode=chartcode,
                            parameter_row=row_number,
                            parameter_text=parameter_text,
                            variable_hints=valid_hints,
                        ),
                        explicit_id=bool(explicit_id),
                        rule_key=_canonical_parameter_rule(parameter_text),
                    ))

                # 一个显式 ID 必须始终表示同一动作、图表和规则。
                signatures_by_parameter_id: dict[
                    str,
                    set[tuple[str, str, str]],
                ] = {}
                for candidate in provisional_parameters:
                    if not candidate.explicit_id:
                        continue
                    entry = candidate.entry
                    signatures_by_parameter_id.setdefault(
                        entry.experience_id,
                        set(),
                    ).add((
                        entry.normalized_operation,
                        normalize_chartcode(entry.chartcode),
                        candidate.rule_key,
                    ))
                conflicted_parameter_ids = {
                    experience_id
                    for experience_id, signatures
                    in signatures_by_parameter_id.items()
                    if len(signatures) > 1
                }
                for experience_id in sorted(conflicted_parameter_ids):
                    issues.append(ExperienceIssue(
                        "warning",
                        "conflicting_parameter_experience_id",
                        (
                            f"参数经验ID“{experience_id}”对应多个动作、"
                            "Chartcode或规则，相关行已禁用"
                        ),
                        _PARAMETER_SHEET,
                    ))

                id_valid_candidates = [
                    candidate
                    for candidate in provisional_parameters
                    if (
                        not candidate.explicit_id
                        or candidate.entry.experience_id
                        not in conflicted_parameter_ids
                    )
                ]

                # 同一动作 + Chartcode 只能存在一套规则；不同规则全部禁用。
                rules_by_action: dict[
                    tuple[str, str],
                    set[str],
                ] = {}
                for candidate in id_valid_candidates:
                    entry = candidate.entry
                    action_key = (
                        entry.normalized_operation,
                        normalize_chartcode(entry.chartcode),
                    )
                    rules_by_action.setdefault(action_key, set()).add(
                        candidate.rule_key
                    )
                conflicted_actions = {
                    action_key
                    for action_key, rules in rules_by_action.items()
                    if len(rules) > 1
                }
                for operation_key, chartcode_key in sorted(
                    conflicted_actions
                ):
                    issues.append(ExperienceIssue(
                        "warning",
                        "conflicting_parameter_experience",
                        (
                            f"动作“{operation_key}”在Chartcode "
                            f"“{chartcode_key}”下存在不同参数经验，"
                            "相关规则已全部禁用"
                        ),
                        _PARAMETER_SHEET,
                    ))

                # 完全相同的动作、图表和规则只保留第一行；经验 ID 的差异
                # 不应让等价规则制造匹配歧义。
                deduplicated: list[_ParameterCandidate] = []
                seen_parameter_rules: set[tuple[str, str, str]] = set()
                for candidate in id_valid_candidates:
                    entry = candidate.entry
                    action_key = (
                        entry.normalized_operation,
                        normalize_chartcode(entry.chartcode),
                    )
                    if action_key in conflicted_actions:
                        continue
                    identity = (*action_key, candidate.rule_key)
                    if identity in seen_parameter_rules:
                        continue
                    seen_parameter_rules.add(identity)
                    deduplicated.append(candidate)

                parameter_entries = [
                    candidate.entry
                    for candidate in deduplicated
                ]

                # 保持原 match() 契约：若参数行能够唯一绑定到 Chartcode
                # 经验，仍将其附着到 ExperienceEntry。独立池并不依赖绑定。
                def compatible_parameters(
                    chart_entry: ExperienceEntry,
                ) -> list[ParameterExperienceEntry]:
                    chartcode_key = normalize_chartcode(
                        chart_entry.chartcode
                    )
                    exact_id = [
                        candidate.entry
                        for candidate in deduplicated
                        if (
                            candidate.entry.experience_id
                            == chart_entry.experience_id
                            and candidate.entry.normalized_operation
                            == chart_entry.normalized_operation
                            and normalize_chartcode(
                                candidate.entry.chartcode
                            ) == chartcode_key
                        )
                    ]
                    if exact_id:
                        return exact_id
                    return [
                        candidate.entry
                        for candidate in deduplicated
                        if (
                            not candidate.explicit_id
                            and candidate.entry.normalized_operation
                            == chart_entry.normalized_operation
                            and normalize_chartcode(
                                candidate.entry.chartcode
                            ) == chartcode_key
                        )
                    ]

                attached_entries: list[ExperienceEntry] = []
                for entry in entries:
                    compatible = compatible_parameters(entry)
                    if len(compatible) != 1:
                        attached_entries.append(entry)
                        continue
                    parameter = compatible[0]
                    attached_entries.append(replace(
                        entry,
                        parameter_row=parameter.parameter_row,
                        parameter_text=parameter.parameter_text,
                        variable_hints=dict(parameter.variable_hints),
                    ))
                entries = attached_entries
        else:
            issues.append(ExperienceIssue(
                "warning", "missing_sheet",
                f"缺少工作表“{_PARAMETER_SHEET}”，Chartcode经验仍可使用",
                _PARAMETER_SHEET,
            ))

        common_entries = _load_common_entries(
            workbook,
            charts,
            chart_lookup,
            issues,
        )
    finally:
        workbook.close()

    index = ExperienceIndex(
        entries,
        parameter_entries=parameter_entries,
        digest=digest,
        source_name=resolved_source_name,
        embed_backend=embed_backend,
    )
    common_index = CommonChartSemanticIndex(
        common_entries,
        embed_backend=embed_backend,
        similarity_threshold=0.70,
    )
    return ExperienceLoadResult(
        index=index,
        issues=tuple(issues),
        digest=digest,
        common_entries=tuple(common_entries),
        common_index=common_index,
    )
